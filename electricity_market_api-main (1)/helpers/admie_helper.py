from datetime import datetime, date
import math
import os
from efc.interfaces.iopenpyxl import OpenpyxlInterface
from openpyxl import load_workbook
import pandas as pd
import pytz
import xlrd
from helpers.date_helper import getTimeDifferncePandasBallancingCap
from helpers.download_helper import downloadJson
from dateutil import parser

from models.file import CustomFile

def getAdmieFiletype(fileType: str, dateFrom: datetime, dateTo: datetime):
    files = downloadJson(f"https://www.admie.gr/getOperationMarketFilewRange?dateStart={dateFrom:%Y-%m-%d}&dateEnd={dateTo:%Y-%m-%d}&FileCategory={fileType}")
    if files is None:
        return []

    for row in files:
        row['file_fromdate'] = datetime.strptime(row['file_fromdate'], '%d.%m.%Y')
        row['file_todate'] = datetime.strptime(row['file_todate'], '%d.%m.%Y')

    return files

def parseAdmieDate(date_str: str) -> datetime:
    for fmt in ('%d/%m/%Y %H:%M', '%d/%m/%y %H:%M'):
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            pass

    # neither worked 
    raise ValueError(f"Time data {date_str!r} does not match any supported format")


def checkStringToInt(date : str):
    try:
        if int(date):
            return True
    except ValueError:
        return False
    

def getDateFromDf(df, market_change_flag: bool):

    if not market_change_flag:
        return datetime.strptime(df.iloc[0,2],'%d/%m/%Y')
    else:
        return df.iloc[0,3]

def getTimeDiffernceCET(date : datetime):
    
    cet_time = date.astimezone(pytz.timezone('Europe/Berlin'))

    # Convert the UTC datetime to CET
    cet_timezone = pytz.timezone('Europe/Berlin')
    #cet_datetime = date.astimezone(cet_timezone)

    # Get the date in CET
    cet_date = cet_time.date()

    # Convert the CET date back to a datetime object (midnight)
    cet_datetime_midnight = datetime.combine(cet_date, datetime.min.time())

    # Convert the CET datetime back to UTC
    cet_datetime_midnight = cet_timezone.localize(cet_datetime_midnight)

    #normalize
    utc_date = pytz.utc.normalize(cet_datetime_midnight)

    cet_date = pytz.utc.normalize(cet_time)

    hour_diff = cet_date-utc_date

    return int(abs(hour_diff.total_seconds()) // 3600) + 1


def getSettlementFromAdmieFileVersion(start_date : datetime ,settlement_timestamp: datetime):

    if settlement_timestamp is None:
        return None
    

    #get difference between start date and settlement offset in days
    day_difference = (settlement_timestamp - start_date).days

    if day_difference >= 0 and day_difference <= 15:
        return "WeekPlus1"
    elif day_difference > 15 and day_difference <= 90:
        return "WeekPlus6"
    else:
        return "WeekPlus52"
    

def readWorkbook(path: str):
    """
    Read either .xls or .xlsx into a pandas DataFrame,
    preserving the first row as column names.
    """
    ext = os.path.splitext(path)[1].lower()
    
    if ext == ".xlsx":
        wb = load_workbook(path, data_only=False)
        sheet = wb.active
        interface = OpenpyxlInterface(wb=wb, use_cache=False)

        return wb, sheet, interface
    elif ext == ".xls":
        wb = xlrd.open_workbook(path)
        sheet = wb.sheet_by_index(0)
        return wb, sheet, None
    else:
        raise ValueError(f"Unsupported extension {ext!r}, must be .xls or .xlsx")
    


def getAdmieEnergyProductDataFromFileType(sheet, time_infered, file_type, file: CustomFile, version: int, results: list, interface : OpenpyxlInterface =None):
    """
    Extracts data from an Admie file based on its type (.xls or .xlsx).
    Args:
        sheet: The sheet object from the workbook.
        time_infered: A list of inferred times for dispatch.
        file_type: The type of the file ('xls' or 'xlsx').
        file: An instance of CustomFile containing metadata about the file.
        version: The version number of the data.
        results: A list to append the extracted data.
        interface: An instance of OpenpyxlInterface for .xlsx files.
    Returns:
        A list of dictionaries containing the extracted data.
        
    """

    if file_type == 'xls':
        for i in range(1, sheet.nrows):
            cells = sheet.row(i)

            # parse the timestamp in your format
            timestamp = parseAdmieDate(cells[0].value)

            # your dispatch time logic
            hour_dispatch   = getTimeDifferncePandasBallancingCap(time_infered[i-1])
            minute_dispatch = int(timestamp.minute) // 15

            mfrrUp = cells[2].value
            mfrrDaUp = cells[3].value
            mfrrDown = cells[4].value
            mfrrDaDown = cells[5].value
            afrrUp = cells[6].value
            afrrDown = cells[7].value

            # handle NaN values
            mfrrUp = 0 if math.isnan(mfrrUp) else float(mfrrUp)
            mfrrDaUp = 0 if math.isnan(mfrrDaUp) else float(mfrrDaUp)
            mfrrDown = 0 if math.isnan(mfrrDown) else float(mfrrDown)
            mfrrDaDown = 0 if math.isnan(mfrrDaDown) else float(mfrrDaDown)
            afrrUp = 0 if math.isnan(afrrUp) else float(afrrUp)
            afrrDown = 0 if math.isnan(afrrDown) else float(afrrDown)

            results.append({
                'DispatchDay': timestamp.date(),
                'DispatchPeriod': 4*(hour_dispatch-1) + minute_dispatch + 1,
                'FileId': file.Id,
                'ZoneId': 1,
                'Version': version,
                'TotalMfrrUp': mfrrUp,
                'TotalMfrrDaUp': mfrrDaUp,
                'TotalMfrrDown': mfrrDown,
                'TotalMfrrDaDown': mfrrDaDown,
                'TotalAfrrUp': afrrUp,
                'TotalAfrrDown': afrrDown,
                "SettlementVersion": getSettlementFromAdmieFileVersion(file.TargetDateFrom, file.PublicationDate),
            })

    elif file_type == 'xlsx':
        for i,row in enumerate(sheet.iter_rows(min_row=2)):
            timestamp =  parseAdmieDate(row[0].value)

            hour_dispatch = hour_dispatch = getTimeDifferncePandasBallancingCap(time_infered[i])
            minute_dispatch = int(timestamp.minute) // 15

            mfrrUp = row[2]
            mfrrDaUp = row[3]
            mfrrDown = row[4]
            mfrrDaDown = row[5]
            afrrUp = row[6]
            afrrDown = row[7]



            # Calculate the formulas for each cell using the OpenpyxlInterface
            mfrrUp = interface.calc_cell(sheet.cell(row=row[2].row, column=row[2].column).coordinate, sheet.title)
            mfrrDaUp = interface.calc_cell(sheet.cell(row=row[3].row, column=row[3].column).coordinate, sheet.title)
            mfrrDown = interface.calc_cell(sheet.cell(row=row[4].row, column=row[4].column).coordinate, sheet.title)
            mfrrDaDown = interface.calc_cell(sheet.cell(row=row[5].row, column=row[5].column).coordinate, sheet.title)
            afrrUp = interface.calc_cell(sheet.cell(row=row[6].row, column=row[6].column).coordinate, sheet.title)
            afrrDown = interface.calc_cell(sheet.cell(row=row[7].row, column=row[7].column).coordinate, sheet.title)

            results.append({
                'DispatchDay': timestamp.date(), 
                'DispatchPeriod': 4*(hour_dispatch-1) + minute_dispatch + 1,
                'FileId': file.Id,
                'ZoneId': 1,
                'Version': version,
                'TotalMfrrUp': 0 if math.isnan(mfrrUp) else float(mfrrUp),
                'TotalMfrrDaUp': 0 if math.isnan(mfrrDaUp) else float(mfrrDaUp),
                'TotalMfrrDown': 0 if math.isnan(mfrrDown) else float(mfrrDown),
                'TotalMfrrDaDown': 0 if math.isnan(mfrrDaDown) else float(mfrrDaDown),
                'TotalAfrrUp': 0 if pd.isna(afrrUp) else float(afrrUp),
                'TotalAfrrDown': 0 if pd.isna(afrrDown) else float(afrrDown),
                "SettlementVersion": getSettlementFromAdmieFileVersion(file.TargetDateFrom, file.PublicationDate),

            })

    return results


def getAdmieCapacityProductDataFromFileType(sheet, time_infered, file_type, file: CustomFile, version: int, results: list, interface : OpenpyxlInterface =None):
    """
    Extracts data from an Admie file based on its type (.xls or .xlsx).
    Args:
        sheet: The sheet object from the workbook.
        time_infered: A list of inferred times for dispatch.
        file_type: The type of the file ('xls' or 'xlsx').
        file: An instance of CustomFile containing metadata about the file.
        version: The version number of the data.
        results: A list to append the extracted data.
        interface: An instance of OpenpyxlInterface for .xlsx files.
    Returns:
        A list of dictionaries containing the extracted data.
        
    """

    if file_type == 'xls':
        for i in range(1, sheet.nrows):
            cells = sheet.row(i)

            # parse the timestamp in your format
            timestamp = parseAdmieDate(cells[0].value)

            # your dispatch time logic
            hour_dispatch   = getTimeDifferncePandasBallancingCap(time_infered[i-1])
            minute_dispatch = int(timestamp.minute) // 15

            raw_fcrUp    = cells[2].value
            raw_fcrDown  = cells[3].value
            raw_afrrUp   = cells[4].value
            raw_afrrDown = cells[5].value
            raw_mfrrUp   = cells[6].value
            raw_mfrrDown = cells[7].value


            fcrUp, fcrDown = (0 if math.isnan(raw_fcrUp) else float(raw_fcrUp),
                            0 if math.isnan(raw_fcrDown) else float(raw_fcrDown))
            afrrUp, afrrDown = (0 if math.isnan(raw_afrrUp) else float(raw_afrrUp),
                                0 if math.isnan(raw_afrrDown) else float(raw_afrrDown))
            mfrrUp, mfrrDown = (0 if math.isnan(raw_mfrrUp) else float(raw_mfrrUp),
                                0 if math.isnan(raw_mfrrDown) else float(raw_mfrrDown))

            results.append({
                'DispatchDay':     timestamp.date(),
                'DispatchPeriod':  4*(hour_dispatch - 1) + minute_dispatch + 1,
                'FileId':          file.Id,
                'ZoneId':          1,
                'Version':         version,
                'TotalFcrUp':      fcrUp,
                'TotalFcrDown':    fcrDown,
                'TotalMfrrUp':     mfrrUp,
                'TotalMfrrDown':   mfrrDown,
                'TotalAfrrUp':     afrrUp,
                'TotalAfrrDown':   afrrDown,
                'TotalRrUp':       0,
                'TotalRrDown':     0,
                "SettlementVersion": getSettlementFromAdmieFileVersion(file.TargetDateFrom, file.PublicationDate),
            })

    elif file_type == 'xlsx':
        for i,row in enumerate(sheet.iter_rows(min_row=2)):
            timestamp =  parseAdmieDate(row[0].value)

            hour_dispatch = hour_dispatch = getTimeDifferncePandasBallancingCap(time_infered[i])
            minute_dispatch = int(timestamp.minute) // 15

            raw_fcrUp = row[2]
            raw_fcrDown = row[3]
            raw_afrrUp = row[4]
            raw_afrrDown = row[5]
            raw_mfrrUp = row[6]
            raw_mfrrDown = row[7]

            
            # Calculate the formulas for each cell using the OpenpyxlInterface
            fcrUp = interface.calc_cell(sheet.cell(row=row[2].row, column=row[2].column).coordinate, sheet.title)
            fcrDown = interface.calc_cell(sheet.cell(row=row[3].row, column=row[3].column).coordinate, sheet.title)
            afrrUp = interface.calc_cell(sheet.cell(row=row[4].row, column=row[4].column).coordinate, sheet.title)
            afrrDown = interface.calc_cell(sheet.cell(row=row[5].row, column=row[5].column).coordinate, sheet.title)
            mfrrUp = interface.calc_cell(sheet.cell(row=row[6].row, column=row[6].column).coordinate, sheet.title)
            mfrrDown = interface.calc_cell(sheet.cell(row=row[7].row, column=row[7].column).coordinate, sheet.title)

            results.append({
                'DispatchDay': timestamp.date(), 
                'DispatchPeriod': 4*(hour_dispatch-1) + minute_dispatch + 1,
                'FileId': file.Id,
                'ZoneId': 1,
                'Version': version,
                'TotalFcrUp': 0 if math.isnan(fcrUp) else float(fcrUp),
                'TotalFcrDown': 0 if math.isnan(fcrDown) else float(fcrDown),
                'TotalMfrrUp': 0 if math.isnan(mfrrUp) else float(mfrrUp),
                'TotalMfrrDown': 0 if math.isnan(mfrrDown) else float(mfrrDown),
                'TotalAfrrUp': 0 if pd.isna(afrrUp) else float(afrrUp),
                'TotalAfrrDown': 0 if pd.isna(afrrDown) else float(afrrDown),
                "SettlementVersion": getSettlementFromAdmieFileVersion(file.TargetDateFrom, file.PublicationDate),

            })

    return results


def fill_dispatch_periods_for_isp_market(dispatch_day, payload: list[dict]) -> list[dict]:
    """
    From 2025-10-01, the Greek ISP market moved from 30-minute to 15-minute dispatch periods.

    The purpose of this function is to make the endpoints backward-compatible. 
    If we must import historical data (before 2025-10-01), we need to 
    convert half-hourly data into quarter-hourly data.

    Normalize a payload so it always has 96 dispatch periods.

    - If dispatch_day < MARKET_CHANGE_DATE:
      Expands each record (assumed 48 half-hourly) into 2 quarter-hourly records.
    - If dispatch_day >= MARKET_CHANGE_DATE:
      Assumes payload is already 96 quarter-hourly records.

    Args:
        dispatch_day (datetime.date or str): The dispatch day.
        payload (list[dict]): Each dict must contain 'dispatchPeriod' and 'dispatchDay'. Other keys are preserved.

    Returns:
        list[dict]: Normalized list of records (96 periods).
    """
    if isinstance(dispatch_day, datetime):
        MARKET_CHANGE_DATE = datetime(2025, 10, 1)

    elif isinstance(dispatch_day, date):
        MARKET_CHANGE_DATE = date(2025, 10, 1)

    # Already in quarter-hour granularity → return unchanged
    if dispatch_day >= MARKET_CHANGE_DATE:
        return payload

    # Pre-change: expand each half-hourly record into 2 quarter-hourly records
    results = []
    for record in payload:
        orig_period = record.get('dispatchPeriod', record.get('DispatchPeriod'))
        day_key = 'dispatchDay' if 'dispatchDay' in record else 'DispatchDay'
        period_key = 'dispatchPeriod' if 'dispatchPeriod' in record else 'DispatchPeriod'
        
        for j in range(1, 3):
            new_record = record.copy()
            new_record[day_key] = dispatch_day
            new_record[period_key] = 2 * (orig_period - 1) + j
            results.append(new_record)

    return results