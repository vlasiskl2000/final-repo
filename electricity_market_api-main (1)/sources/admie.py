import numpy as np
from helpers.download_helper import *
from helpers.date_helper import *
from helpers.metadata_helper import *
from helpers.admie_helper import *
from helpers.log_helper import logException
from helpers.file_helper import *
from helpers.scada_helper import *

from typing import List
from models.admie_models import BalancingEnergyData, EnergySurplusData
from models.file import CustomFile
from models.metadata import *
from interface.envelope import Envelope

import pandas as pd
from datetime import datetime, timedelta

import math
import os
from openpyxl import load_workbook

from itertools import groupby

market_change_date = datetime(2025, 10, 1)

## ISP Documents Parsing 
def getUnitAvailabilities(dateFrom, dateTo, isp=None, version=None):
    files = getISPAvailabilityFiles(dateFrom, dateTo, isp, version)

    for file in files:
        yield getUnitAvailabilitiesFromFile(file)

def getUnitAvailabilitiesFromFile(file: CustomFile):
    results = []

    UnitProductionColumns = ["EntityName", "PublishedCapacity", "EstimatedCapacity", "Reason"]

    dispatchDate = datetime.strptime(file.FileName.split('_')[0], '%Y%m%d')
    version = file.PublicationDate

    path = getFileNameFromInfo(file)
    if not os.path.exists(path):
        message = f'{path} does not exist'
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        UnitAvailabilities = pd.read_excel(path, skiprows = 4, usecols = range(1,5), header = None)
    except Exception as e:
        message = f'Failed to read {path}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)
    
    try:
        UnitAvailabilities.columns = UnitProductionColumns
        UnitAvailabilities = UnitAvailabilities.fillna('')

        for i in range(len(UnitAvailabilities)):
            dict = {}

            row = UnitAvailabilities.iloc[i]

            dict["DispatchDay"] = dispatchDate
            dict["EntityName"] = row['EntityName']
            dict["PublishedCapacity"] = float(row['PublishedCapacity'] if row['PublishedCapacity'] != '' else 0)
            dict["EstimatedCapacity"] = float(row['EstimatedCapacity'] if row['EstimatedCapacity'] != '' else 0)
            dict["Reason"] = row['Reason']
            dict["Version"] = version
            #dict["VersionDescription"] = versionDescription
            dict["FileId"] = file.Id

            results.append(dict)

        return envelope.getSuccessEnvelope(getParsingMetadataPayload(results, getMetadata(True, file.Id)))
    except Exception as e:
        message = f'Failed to parse {path}: {e}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)
    
def getISPEnergyOffers(dateFrom, dateTo):
    files = getFileUrlsFromApi("ISPEnergyOffers", dateFrom, dateTo)

    for file in files:
        yield getISPEnergyOffersFromFile(file)

def getISPEnergyOffersFromFile(file: CustomFile):
    results = []

    path = getFileNameFromInfo(file)
    if not os.path.exists(path):
        message = f'{path} does not exist'
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        ISPEnergyOffersExcel = pd.ExcelFile(path)
        sheets = ISPEnergyOffersExcel.sheet_names
    except Exception as e:
        message = f'Failed to read {path}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    #version = datetime.strptime(file["PublicationDate"], '%d.%m.%Y %H:%M')
    for sheet in range(len(sheets)):
        try:
            ISPEnergyOffers = pd.read_excel(path, sheet_name=sheet)
        except Exception as e:
            message = f'Failed to read {path}'
            logException(e)
            return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

        try:
            ISPEnergyOffers = ISPEnergyOffers.fillna('')
            
            # Check if target date is before October 1, 2025 to determine if row duplication is needed
            target_date = file.TargetDateFrom
            cutoff_date = datetime(2025, 10, 1)

            duplicate_rows = target_date < cutoff_date
            row_count = 2 if duplicate_rows else 1
            
            for i in range(len(ISPEnergyOffers)):
                row = ISPEnergyOffers.iloc[i]
                if row["ID_PERIOD"] == '': 
                    continue
                dispatchDate = row["ID_PERIOD"].to_pydatetime()

                for j in range(row_count):
                    dict = {}
                    dict["DispatchDay"] = dispatchDate.strftime('%Y-%m-%d')
                    if duplicate_rows:
                        dict["DispatchPeriod"] = 2 * ((dispatchDate.hour*2 + 1) + (1 if dispatchDate.minute == 30 else 0)) - (1 if j == 0 else 0)
                    else:
                        # For 15-minute data, use the actual dispatch period from the time
                        dict["DispatchPeriod"] = (dispatchDate.hour * 4) + (dispatchDate.minute // 15) + 1
                    dict["Quantity"] = 0 if row['QUANTITY_MW'] == '' else np.round(float(row['QUANTITY_MW']),3)
                    dict["Price"] = 0 if row['PRICE'] == '' else np.round(float(row['PRICE']),3)
                    dict["Step"] =  int(row['SEG']) if "SEG" in row else None
                    dict["Up"] = row['DIR'] == 'Up'
                    dict["MinQuantity"] = 0 if row['MINQUANTITY'] == '' else np.round(float(row['MINQUANTITY']),3)
                    dict["FileId"] = file.Id
                    results.append(dict)
        
        except Exception as e:
            message = f'Failed to parse {path}: {e}'
            logException(e)
            return envelope.getFailedEnvelope(getParsingMetadataPayload(results, getMetadata(False, file.Id, message)), None)
    
    return envelope.getSuccessEnvelope(getParsingMetadataPayload(results, getMetadata(True, file.Id)))

def getISPCapacityOffers(dateFrom, dateTo):
    files = getFileUrlsFromApi("ISPCapacityOffers", dateFrom, dateTo)

    for file in files:
        yield getISPCapacityOffersFromFile(file)

def getISPCapacityOffersFromFile(file: CustomFile):
    results = []

    path = getFileNameFromInfo(file)
    if not os.path.exists(path):
        message = f'{path} does not exist'
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        ISPEnergyOffersExcel = pd.ExcelFile(path)
        sheets = ISPEnergyOffersExcel.sheet_names
    except Exception as e:
        message = f'Failed to read {path}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        ISPEnergyOffersList = []
        for i in range(len(sheets)):
            ISPEnergyOffersList.append(pd.read_excel(path, sheets[i]))
        
        ISPEnergyOffers = pd.concat(ISPEnergyOffersList)
        ISPEnergyOffers = ISPEnergyOffers.fillna('')
        
        # Check if target date is before October 1, 2025 to determine if row duplication is needed
        target_date = file.TargetDateFrom
        cutoff_date = datetime(2025, 10, 1)

        duplicate_rows = target_date < cutoff_date
        row_count = 2 if duplicate_rows else 1
        
        for i in range(len(ISPEnergyOffers)):
            row = ISPEnergyOffers.iloc[i]
            if row["ID_PERIOD"] == '': 
                continue
            
            for j in range(row_count):
                dict = {}

                dispatchDate = row["ID_PERIOD"].to_pydatetime()
                dict["DispatchDay"] = dispatchDate.strftime('%Y-%m-%d')
                if duplicate_rows:
                    dict["DispatchPeriod"] = 2 * ((dispatchDate.hour*2 + 1) + (1 if dispatchDate.minute == 30 else 0)) - (1 if j == 0 else 0)
                else:
                    # For 15-minute data, use the actual dispatch period from the time
                    dict["DispatchPeriod"] = (dispatchDate.hour * 4) + (dispatchDate.minute // 15) + 1
                dict["Quantity"] = 0 if row['QUANTITY_MW'] == '' else float(row['QUANTITY_MW'])
                dict["Price"] = 0 if row['PRICE'] == '' else float(row['PRICE'])
                dict["Step"] =  int(row['SEG']) if "SEG" in row else None
                dict["Up"] = row['DIR'] == 'Up'
                dict["ReserveType"] = RESERVE_TYPE[row['SERVICETYPE']]
                dict["MinQuantity"] = 0 if row['MINQUANTITY'] == '' else float(row['MINQUANTITY'])
                dict["FileId"] = file.Id

                results.append(dict)

    except Exception as e:
        message = f'Failed to parse {path}: {e}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    return envelope.getSuccessEnvelope(getParsingMetadataPayload(results, getMetadata(True, file.Id)))


## ISP Requirements
def getMandatoryHydroFromISPRequirements(dateFrom: str, dateTo : str, isp=None, generateLinks = False):
    if generateLinks:
        files = getISPRequirementsGeneratedLinks(dateFrom, dateTo)
    else:
        files = getISPFilesWithPrefix("Requirements", dateFrom, dateTo, isp, None, 4)

    for file in files:
        yield getMandatoryHydroFromFile(file)

def getMandatoryHydroFromFile(file: CustomFile):
    results = []
    version = file.PublicationDate

    path = getFileNameFromInfo(file)
    if not os.path.exists(path):
        message = f'{path} does not exist'
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        ISPRequirements = pd.read_excel(path, skiprows = [0], header = 0)
        ISPRequirements = ISPRequirements.fillna('')
    except Exception as e:
        message = f'Failed to read {path}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)
    
    try:
        first_column = ISPRequirements.columns[0]

        MandatoryHydro = []
        
        index_mandatory_hydro = ISPRequirements.index[ISPRequirements[first_column] == 'Mandatory Hydro'].values[0]
        index_commissioning = ISPRequirements.index[ISPRequirements[first_column] == 'Commissioning'].values[0]

        dispatchDate = file.TargetDateFrom
        getISPResultsDataFromRange(ISPRequirements, index_mandatory_hydro + 1 , index_commissioning, None, MandatoryHydro, 1, 2  , 1, dispatchDateStr = dispatchDate )
        
        for row in MandatoryHydro:
            results.append({
                'DispatchDay': dispatchDate,
                'DispatchPeriod': row[1],
                'EntityName': row[2],
                'Generation': row[3],
                'Version': version,
                'FileId': file.Id
                #'VersionDescription': versionDescription
            })

        results = fill_dispatch_periods_for_isp_market(dispatchDate, results)

        return envelope.getSuccessEnvelope(getParsingMetadataPayload(results, getMetadata(True, file.Id)))
    except Exception as e:
        message = f'Failed to parse {path}: {e}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

def getCommissiongFromISPRequirements(dateFrom: str, dateTo : str, isp=None, generateLinks = False):
    if generateLinks:
        files = getISPRequirementsGeneratedLinks(dateFrom, dateTo)
    else:
        files = getISPFilesWithPrefix("Requirements", dateFrom, dateTo, isp, None, 4)

    for file in files:
        yield getCommissiongFromFile(file)

def getCommissiongFromFile(file: CustomFile):
    results = []
    version = file.PublicationDate

    path = getFileNameFromInfo(file)
    if not os.path.exists(path):
        message = f'{path} does not exist'
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        ISPRequirements = pd.read_excel(path, skiprows = [0], header = 0)
        ISPRequirements = ISPRequirements.fillna('')
    except Exception as e:
        message = f'Failed to read {path}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)
    
    try:
        first_column = ISPRequirements.columns[0]

        Commissioning = []
        
        index_commissioning = ISPRequirements.index[ISPRequirements[first_column] == 'Commissioning'].values[0]
        index_reserve_requirements = ISPRequirements.index[ISPRequirements[first_column] == 'Reserve Requirements'].values[0]

        dispatchDate = file.TargetDateFrom
        getISPResultsDataFromRange(ISPRequirements, index_commissioning + 1, index_reserve_requirements, None, Commissioning, 1, 2, 1, dispatchDateStr = dispatchDate) 
        
        for row in Commissioning:
            results.append({
                'DispatchDay': dispatchDate,
                'DispatchPeriod': row[1],
                'EntityName': row[2],
                'Value': row[3],
                'Version': version,
                'FileId': file.Id
                #'VersionDescription': versionDescription
            })

        results = fill_dispatch_periods_for_isp_market(dispatchDate, results)

        return envelope.getSuccessEnvelope(getParsingMetadataPayload(results, getMetadata(True, file.Id)))
    except Exception as e:
        message = f'Failed to parse {path}: {e}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

def getIspSystemLossesFromRequirements(dateFrom: str, dateTo : str, isp=None, generateLinks = False):
    if generateLinks:
        files = getISPRequirementsGeneratedLinks(dateFrom, dateTo)
    else:
        files = getISPFilesWithPrefix("Requirements", dateFrom, dateTo, isp, None, 4)

    for file in files:
        yield getIspSystemLossesFromFile(file)

def getIspSystemLossesFromFile(file: CustomFile):
    results = []
    version = file.PublicationDate

    path = getFileNameFromInfo(file)
    if not os.path.exists(path):
        message = f'{path} does not exist'
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        ISPRequirements = pd.read_excel(path, skiprows = [0], header = None)
        ISPRequirements = ISPRequirements.fillna('')
    except Exception as e:
        message = f'Failed to read {path}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)
    
    try:
        first_column = ISPRequirements.columns[0]
        
        first_hours = ISPRequirements.iloc[0,2]
        dispatchPeriod = getDispatchPeriodFromHourString(first_hours)

        index_system_losses = ISPRequirements.index[(ISPRequirements[first_column] == 'Non-Dispatcheble Losses') | (ISPRequirements[first_column] == 'Non-Dispatchable Losses')].values[0]

        for t in range(1, ISPRequirements.shape[1] - 2, 1):
            value = ISPRequirements[t+1][index_system_losses]
            results.append({'dispatchDay': file.TargetDateFrom,
                'dispatchPeriod': dispatchPeriod,
                'version': version,
                'fileId': file.Id,
                'zoneId': 1,
                'value': value})
            dispatchPeriod += 1

        results = fill_dispatch_periods_for_isp_market(file.TargetDateFrom, results)

        return envelope.getSuccessEnvelope(getParsingMetadataPayload(results, getMetadata(True, file.Id)))
    except Exception as e:
        message = f'Failed to parse {path}: {e}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

def getIspResForecastsFromISPRequirementsFile(file: CustomFile):
    results = []
    version = file.PublicationDate

    path = getFileNameFromInfo(file)
    if not os.path.exists(path):
        message = f'{path} does not exist'
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        ISPRequirements = pd.read_excel(path, skiprows = [0], header = None)
        ISPRequirements = ISPRequirements.fillna('')
    except Exception as e:
        message = f'Failed to read {path}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)
    
    try:
        first_column = ISPRequirements.columns[0]
        
        first_hours = ISPRequirements.iloc[0,2]
        dispatchPeriod = getDispatchPeriodFromHourString(first_hours)


        index_res = ISPRequirements.index[(ISPRequirements[first_column] == 'Non-Dispatchable RES') | (ISPRequirements[first_column] == 'Non-Dispatcheble RES')].values[0]

        for t in range(1, ISPRequirements.shape[1] - 2, 1):
            value = ISPRequirements[t+1][index_res]
            results.append({'dispatchDay': file.TargetDateFrom,
                'dispatchPeriod': dispatchPeriod,
                'version': version,
                'fileId': file.Id,
                'zoneId': 1,
                'resForecast': value})
            dispatchPeriod += 1

        results = fill_dispatch_periods_for_isp_market(file.TargetDateFrom, results)

        return envelope.getSuccessEnvelope(getParsingMetadataPayload(results, getMetadata(True, file.Id)))
    except Exception as e:
        message = f'Failed to parse {path}: {e}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

def getIspLoadForecastsFromISPRequirementsFile(file: CustomFile):
    results = []
    version = file.PublicationDate

    path = getFileNameFromInfo(file)
    if not os.path.exists(path):
        message = f'{path} does not exist'
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        ISPRequirements = pd.read_excel(path, skiprows = [0], header = None)
        ISPRequirements = ISPRequirements.fillna('')
    except Exception as e:
        message = f'Failed to read {path}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)
    
    try:
        first_column = ISPRequirements.columns[0]
        
        index_res = ISPRequirements.index[(ISPRequirements[first_column] == 'Non-Dispatchable Load') | (ISPRequirements[first_column] == 'Non-Dispatcheble Load')].values[0]

        dispatchPeriod = getDispatchPeriodFromHourString(file.TargetDateFrom.strftime('%H:%M'))

        for t in range(1, ISPRequirements.shape[1] - 2, 1):
            value = ISPRequirements[t+1][index_res]
            results.append({'dispatchDay': file.TargetDateFrom,
                'dispatchPeriod': dispatchPeriod,
                'version': version,
                'fileId': file.Id,
                'zoneId': 1,
                'loadForecast': value})
            dispatchPeriod += 1

        results = fill_dispatch_periods_for_isp_market(file.TargetDateFrom, results)

        return envelope.getSuccessEnvelope(getParsingMetadataPayload(results, getMetadata(True, file.Id)))
    except Exception as e:
        message = f'Failed to parse {path}: {e}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

def getIspCommissioningScheduleFromRequirements(dateFrom: str, dateTo : str, isp=None, generateLinks = False):
    if generateLinks:
        files = getISPRequirementsGeneratedLinks(dateFrom, dateTo)
    else:
        files = getISPFilesWithPrefix("Requirements", dateFrom, dateTo, isp, None, 4)

    for file in files:
        yield getIspCommissioningScheduleFromFile(file)

def getIspCommissioningScheduleFromFile(file: CustomFile):
    results = []
    version = file.PublicationDate

    path = getFileNameFromInfo(file)
    if not os.path.exists(path):
        message = f'{path} does not exist'
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        ISPRequirements = pd.read_excel(path, skiprows = [0], header = None)
        ISPRequirements = ISPRequirements.fillna('')
    except Exception as e:
        message = f'Failed to read {path}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)
    
    try:
        first_column = ISPRequirements.columns[0]
        
        index = ISPRequirements.index[ISPRequirements[first_column] == 'Commissioning'].values[-1]

        for t in range(1, ISPRequirements.shape[1]-2, 1):
            value = ISPRequirements[t+1][index]
            results.append({'dispatchDay': file.TargetDateFrom,
                'dispatchPeriod': t,
                'version': version,
                'fileId': file.Id,
                'zoneId': 1,
                'value': value})

        results = fill_dispatch_periods_for_isp_market(file.TargetDateFrom, results)

        return envelope.getSuccessEnvelope(getParsingMetadataPayload(results, getMetadata(True, file.Id)))
    except Exception as e:
        message = f'Failed to parse {path}: {e}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

def getReserveRequirementsFromISPRequirements(dateFrom: datetime, dateTo: datetime, isp=None, generateLinks = False):
    if generateLinks:
        files = getISPRequirementsGeneratedLinks(dateFrom, dateTo)
    else:
        files = getISPFilesWithPrefix("Requirements", dateFrom, dateTo, isp, None, 4)

    for file in files:
        yield getReserveRequirementsFromFile(file)

def getReserveRequirementsFromFile(file: CustomFile):
    results = []
    version = file.PublicationDate

    path = getFileNameFromInfo(file)
    if not os.path.exists(path):
        message = f'{path} does not exist'
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        ISPRequirements = pd.read_excel(path, skiprows = [0], header = 0)
        ISPRequirements = ISPRequirements.fillna('')
    except Exception as e:
        message = f'Failed to read {path}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)
    
    try:
        first_column = ISPRequirements.columns[0]
        
        ReserveRequirementsUp = []
        ReserveRequirementsDown = []
        
        index_rr_up = ISPRequirements.index[ISPRequirements[first_column] == 'Up'].values[0]
        index_rr_down = ISPRequirements.index[ISPRequirements[first_column] == 'Down'].values[0]

        dispatchDate = file.TargetDateFrom
        getISPResultsDataFromRange(ISPRequirements, index_rr_up, index_rr_down, 'Up', ReserveRequirementsUp,  start_col = 2, entityName_col = 1, dispatchDateStr = dispatchDate)
        getISPResultsDataFromRange(ISPRequirements, index_rr_down, len(ISPRequirements.index) - 1, 'Down', ReserveRequirementsDown, start_col = 2, entityName_col = 1, dispatchDateStr = dispatchDate)

        ReserveRequirements = ReserveRequirementsUp + ReserveRequirementsDown

        key = lambda x: {
            'dispatchDay': x[0],
            'dispatchPeriod': x[1]
        }

        ReserveRequirements.sort(key=lambda x: x[1])
        ReserveRequirements.sort(key=lambda x: x[0])
        for k,v in groupby(ReserveRequirements,key=key):
            values = list(v)

            fcrUp = list(filter(lambda x: x[2] == 'Up' and x[3] == 'FCR', values))
            fcrDown = list(filter(lambda x: x[2] == 'Down' and x[3] == 'FCR', values))
            afrrUp = list(filter(lambda x: x[2] == 'Up' and x[3] == 'aFRR', values))
            afrrDown = list(filter(lambda x: x[2] == 'Down' and x[3] == 'aFRR', values))
            mfrrUp = list(filter(lambda x: x[2] == 'Up' and x[3] == 'mFRR', values))
            mfrrDown = list(filter(lambda x: x[2] == 'Down' and x[3] == 'mFRR', values))
            rrUp = list(filter(lambda x: x[2] == 'Up' and x[3] == 'RR', values))
            rrDown = list(filter(lambda x: x[2] == 'Down' and x[3] == 'RR', values))
            srrUp = list(filter(lambda x: x[2] == 'Up' and x[3] == 'Spinning RR', values))
            srrDown = list(filter(lambda x: x[2] == 'Down' and x[3] == 'Spinning RR', values))

            row = {
                'DispatchDay': dispatchDate,
                'DispatchPeriod': k['dispatchPeriod'],
                'ZoneId': 1,
                'Version': version,
                'FileId': file.Id
            }

            if len(fcrUp) > 0:
                row['Fcrup'] = fcrUp[0][4]
            if len(fcrUp) > 0:
                row['Fcrdown'] = fcrDown[0][4]
            if len(afrrUp) > 0:
                row['AFrrup'] = afrrUp[0][4]
            if len(afrrDown) > 0:
                row['AFrrdown'] = afrrDown[0][4]
            if len(mfrrUp) > 0:
                row['MFrrup'] = mfrrUp[0][4]
            if len(mfrrDown) > 0:
                row['MFrrdown'] = mfrrDown[0][4]
            if len(rrUp) > 0:
                row['Rrup'] = rrUp[0][4]
            if len(rrDown) > 0:
                row['Rrdown'] = rrDown[0][4]
            if len(srrUp) > 0:
                row['SRrup'] = srrUp[0][4]
            if len(srrDown) > 0:
                row['SRrdown'] = srrDown[0][4]

            results.append(row)

        results = fill_dispatch_periods_for_isp_market(dispatchDate, results)

        return envelope.getSuccessEnvelope(getParsingMetadataPayload(results, getMetadata(True, file.Id)))
    except Exception as e:
        message = f'Failed to parse {path}: {e}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)


##


## ISP Deviations
def getIspDeviationswithSuffix(dateFrom : str, dateTo : str):
    files = getFileUrlsFromApi("Devit", dateFrom, dateTo) 
    
    for file in files:
        yield getIspItalyDevFromFile(file)

def getIspItalyDevFromFile(file: CustomFile):
    results = []
    version = file.PublicationDate

    path = getFileNameFromInfo(file)
    if not os.path.exists(path):
        message = f'{path} does not exist'
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        ItalyDev = pd.read_excel(path, header = None)
        ItalyDev = ItalyDev.fillna('')
    except Exception as e:
        message = f'Failed to read {path}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)
    
    try:
        for row in ItalyDev.values[1:]:
            date = row[0]
            for i in range(1, 25):
                if i >= len(row):
                    break
                for j in range(1, 5):
                    results.append({
                        'DispatchDay': date,
                        'DispatchPeriod': 4*(i- 1) + j,
                        'Value': float(row[i]),
                        'Version': version
                    })

        return envelope.getSuccessEnvelope(getParsingMetadataPayload(results, getMetadata(True, file.Id)))
    except Exception as e:
        message = f'Failed to parse {path}: {e}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

def getIspNorthDev(dateFrom : str, dateTo : str):
    files = getFileUrlsFromApi("Devnor", dateFrom, dateTo) 
    
    for file in files:
        yield getIspNorthDevFromFile(file)

def getIspNorthDevFromFile(file: CustomFile):
    results = []
    version = file.PublicationDate

    path = getFileNameFromInfo(file)
    if not os.path.exists(path):
        message = f'{path} does not exist'
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        NorthDev = pd.read_excel(path, header = None)
        NorthDev = NorthDev.fillna('')
    except Exception as e:
        message = f'Failed to read {path}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)
    
    try:
        for row in NorthDev.values[1:]:
            date = row[0]
            for i in range(1, 25):
                if i >= len(row):
                    break
                for j in range(1, 5):
                    results.append({
                        'DispatchDay': date,
                        'DispatchPeriod': 4*(i- 1) + j,
                        'Value': float(row[i]),
                        'Version': version
                    })

        return envelope.getSuccessEnvelope(getParsingMetadataPayload(results, getMetadata(True, file.Id)))
    except Exception as e:
        message = f'Failed to parse {path}: {e}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

## ISP Results

def getReserveRequirementsFromISPResults(dateFrom: datetime, dateTo: datetime):
    files = getFileUrlsFromApi("AdhocISPResults", dateFrom, dateTo) + getFileUrlsFromApi("ISP1ISPResults", dateFrom, dateTo) + getFileUrlsFromApi("ISP2ISPResults", dateFrom, dateTo) + getFileUrlsFromApi("ISP3ISPResults", dateFrom, dateTo)

    for file in files:
        yield getReserveRequirementsFromISPResultsFromFile(file)

def getReserveRequirementsFromISPResultsFromFile(file: CustomFile):

    results = []

    path = getFileNameFromInfo(file)
    if not os.path.exists(path):
        message = f'{path} does not exist'
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        ISP_DataFrame = pd.read_excel(path, skiprows = 1, sheet_name=0)
    except Exception as e:
        message = f'Failed to read {path}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)
    
    try:
        version = file.PublicationDate
        first_column = ISP_DataFrame.columns[0]

        index_price_fcr_up = ISP_DataFrame.index[ISP_DataFrame[first_column] == 'FCR Up'].values[-1]
        index_price_fcr_dn = ISP_DataFrame.index[ISP_DataFrame[first_column] == 'FCR Down'].values[-1]
        index_price_afrr_up = ISP_DataFrame.index[ISP_DataFrame[first_column] == 'aFRR Up'].values[-1]
        index_price_afrr_dn = ISP_DataFrame.index[ISP_DataFrame[first_column] == 'aFRR Down'].values[-1]
        index_price_mfrr_up = ISP_DataFrame.index[ISP_DataFrame[first_column] == 'mFRR Up'].values[-1]
        index_price_mfrr_dn = ISP_DataFrame.index[ISP_DataFrame[first_column] == 'mFRR Down'].values[-1]
        index_price_rr_up = ISP_DataFrame.index[ISP_DataFrame[first_column] == 'RR Up'].values[-1]
        index_price_rr_dn = ISP_DataFrame.index[ISP_DataFrame[first_column] == 'RR Down'].values[-1]
        index_price_srr_up = ISP_DataFrame.index[ISP_DataFrame[first_column] == 'Spinning RR Up'].values[-1]
        index_price_srr_dn = ISP_DataFrame.index[ISP_DataFrame[first_column] == 'Spinning RR Down'].values[-1]
        
        dispatchDate = ISP_DataFrame.columns[1].date()
        first_hours = ISP_DataFrame.columns[1]
        first_hours_pd = pd.Timestamp(first_hours)
        period = getTimeDifferncePandas(first_hours_pd, seconds_granularity=900)

        for i in range(1, len(ISP_DataFrame.columns) - 1):
            
            if type(ISP_DataFrame.columns[i]) is not datetime:
                if not checkDateToString(ISP_DataFrame.columns[i]):
                    continue

            fcr_up_data = ISP_DataFrame.iloc[index_price_fcr_up, :].values
            fcr_down_data = ISP_DataFrame.iloc[index_price_fcr_dn, :].values
            afrr_up_data = ISP_DataFrame.iloc[index_price_afrr_up, :].values
            afrr_down_data = ISP_DataFrame.iloc[index_price_afrr_dn, :].values
            mfrr_up_data = ISP_DataFrame.iloc[index_price_mfrr_up, :].values
            mfrr_down_data = ISP_DataFrame.iloc[index_price_mfrr_dn, :].values
            rr_up_data = ISP_DataFrame.iloc[index_price_rr_up, :].values
            rr_down_data = ISP_DataFrame.iloc[index_price_rr_dn, :].values
            srr_up_data = ISP_DataFrame.iloc[index_price_srr_up, :].values
            srr_down_data = ISP_DataFrame.iloc[index_price_srr_dn, :].values

            
            results.append({
                'DispatchDay': dispatchDate,
                'DispatchPeriod': period,
                'Version': version,
                'ZoneId': 1,
                'FileId': file.Id,
                'Fcrup': fcr_up_data[i],
                'Fcrdown': fcr_down_data[i],
                'AFrrup': afrr_up_data[i],
                'AFrrdown': afrr_down_data[i],
                'MFrrup': mfrr_up_data[i],
                'MFrrdown': mfrr_down_data[i],
                'Rrup': rr_up_data[i],
                'Rrdown': rr_down_data[i],
                'SRrup': srr_up_data[i],
                'SRrdown': srr_down_data[i],
            })
            period += 1 
        
        results = fill_dispatch_periods_for_isp_market(dispatchDate, results)
            
        return envelope.getSuccessEnvelope(getParsingMetadataPayload(results, getMetadata(True, file.Id)))
    except Exception as e:
        message = f'Failed to parse {path}: {e}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

def getActivatedEnergyDataFromISPResults(dateFrom: str, dateTo : str, isp: int = None, adhoc: bool =None):
    files = getISPResultsFiles(dateFrom, dateTo, isp, adhoc)
    
    for file in files:
        yield getIspActivatedBeActualFromFile(file)

def getIspActivatedBeActualFromFile(file: CustomFile):

    results = [] 
    version = file.PublicationDate

    path = getFileNameFromInfo(file)
    if not os.path.exists(path):
        message = f'{path} does not exist'
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        ISP_DataFrame = pd.read_excel(path, skiprows = 1, sheet_name=1)
    except Exception as e:
        message = f'Failed to read {path}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)
    
    try:
        first_column = ISP_DataFrame.columns[0]

        index_thermal_units = 0
        index_hydro_units = ISP_DataFrame.index[ISP_DataFrame[first_column] == 'Hydro Units'].values[0]
        index_res_units = ISP_DataFrame.index[ISP_DataFrame[first_column] == 'RES Units'].values[0]
        index_dispatchable_load = ISP_DataFrame.index[ISP_DataFrame[first_column] == 'Dispatchable Load'].values[0]

        ActivatedEnergy = []
        
        getISPResultsDataFromRange(ISP_DataFrame, index_thermal_units, len(ISP_DataFrame.values), "Thermal", ActivatedEnergy, 1)
        getISPResultsDataFromRange(ISP_DataFrame, index_hydro_units + 1, len(ISP_DataFrame.values), "Hydro", ActivatedEnergy, 1)
        getISPResultsDataFromRange(ISP_DataFrame, index_res_units + 1, len(ISP_DataFrame.values), "RES", ActivatedEnergy, 1)
        getISPResultsDataFromRange(ISP_DataFrame, index_dispatchable_load + 1, len(ISP_DataFrame.values), "Dispatchable", ActivatedEnergy, 1)
    
        demand_response = ISP_DataFrame.index[ISP_DataFrame[first_column] == 'Demand Response']
        if len(demand_response.values) > 0:
            index_demand_response = demand_response.values[0]
            getISPResultsDataFromRange(ISP_DataFrame, index_demand_response + 1, len(ISP_DataFrame.values), "Demand Response", ActivatedEnergy, 1)

        for row in ActivatedEnergy:
            #for j in range(1,3):
            if row[0] == (None,None):
                
                results.append({
                'DispatchDay': datetime.strptime(previous_date, '%Y-%m-%d'),
                'DispatchPeriod': row[1],
                'EntityName': row[3],
                'ActivatedBalancingEnergyUp': row[4] if row[4] > 0 else 0,
                'ActivatedBalancingEnergyDown': -row[4] if row[4] < 0 else 0,
                'Version': version,
                #'VersionDescription': versionDescription
            })
            else:
                results.append({
                'DispatchDay': datetime.strptime(row[0], '%Y-%m-%d'),
                'DispatchPeriod': row[1],
                'EntityName': row[3],
                'ActivatedBalancingEnergyUp': row[4] if row[4] > 0 else 0,
                'ActivatedBalancingEnergyDown': -row[4] if row[4] < 0 else 0,
                'Version': version,
                #'VersionDescription': versionDescription
            })
                previous_date = row[0]

        results = fill_dispatch_periods_for_isp_market(file.TargetDateFrom, results)

        return envelope.getSuccessEnvelope(getParsingMetadataPayload(results, getMetadata(True, file.Id)))
    except Exception as e:
        message = f'Failed to parse {path}: {e}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

def getIspReserveAwardsAndPricesActual(dateFrom: datetime, dateTo: datetime, isp: int = None, adhoc: bool =None):
    files = getISPResultsFiles(dateFrom, dateTo,isp, adhoc)

    for file in files:
        yield getIspReserveAwardsAndPricesActualFromFile(file)

def getIspReserveAwardsAndPricesActualFromFile(file: CustomFile):
    results = []
    AwardsData = []
    version = file.PublicationDate

    path = getFileNameFromInfo(file)
    if not os.path.exists(path):
        message = f'{path} does not exist'
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        xls = pd.ExcelFile(path)
    except Exception as e:
        message = f'Failed to read {path}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)
    
    try:
        def parseSheet(reserve: str, sheet: int):
            ISPResults = pd.read_excel(xls, skiprows = [0], sheet_name=sheet)
        
            first_column = ISPResults.columns[0]

            index_thermal_units = ISPResults.index[ISPResults[first_column] == 'Thermal Unit'].values[0]
            index_hydro_units = ISPResults.index[ISPResults[first_column] == 'Hydro Unit'].values[0]

            getISPResultsDataFromRange(ISPResults, index_thermal_units + 1, len(ISPResults.index), reserve, AwardsData, 2, split_name = True, skipLastColumn=False)
            getISPResultsDataFromRange(ISPResults, index_hydro_units + 1, len(ISPResults.index), reserve, AwardsData, 2, split_name = True, skipLastColumn=False)

            dispatchable_units = ISPResults.index[ISPResults[first_column] == 'Dispatchable Load Unit']
            if len(dispatchable_units.values) > 0:
                index_dispatchable_load = dispatchable_units.values[0]
                getISPResultsDataFromRange(ISPResults, index_dispatchable_load + 1, len(ISPResults.index), reserve, AwardsData, 2, split_name = True, skipLastColumn=False, entitySuffix='_PUMP')

            res_units = ISPResults.index[ISPResults[first_column] == 'RES Unit']
            if len(res_units.values) > 0:
                index_res_units = res_units.values[0]
                getISPResultsDataFromRange(ISPResults, index_res_units + 1, len(ISPResults.index), reserve, AwardsData, 2, split_name = True, skipLastColumn=False)
            
            demand_response_units = ISPResults.index[ISPResults[first_column] == 'Demand Response Unit']
            if len(demand_response_units.values) > 0:
                index_demand_response_units = demand_response_units.values[0]
                getISPResultsDataFromRange(ISPResults, index_demand_response_units + 1, len(ISPResults.index), reserve, AwardsData, 2, split_name = True, skipLastColumn=False)
            
        parseSheet('FCR', 3)
        parseSheet('aFRR', 4)
        parseSheet('mFRR', 5)
        parseSheet('RR', 6)

        key = lambda x: {
            'dispatchDay': x[0],
            'dispatchPeriod': x[1],
            'entityName': x[3]
        }
        
        AwardsData.sort(key=lambda x: x[3])
        AwardsData.sort(key=lambda x: x[1])
        AwardsData.sort(key=lambda x: x[0])
        
        for k,v in groupby(AwardsData,key=key):
            values = list(v)

            fcrUp = list(filter(lambda x: x[4] == 'Up' and x[2] == 'FCR', values))
            fcrDown = list(filter(lambda x: x[4] == 'Down' and x[2] == 'FCR', values))
            afrrUp = list(filter(lambda x: x[4] == 'Up' and x[2] == 'aFRR', values))
            afrrDown = list(filter(lambda x: x[4] == 'Down' and x[2] == 'aFRR', values))
            mfrrUp = list(filter(lambda x: x[4] == 'Up' and x[2] == 'mFRR', values))
            mfrrDown = list(filter(lambda x: x[4] == 'Down' and x[2] == 'mFRR', values))
            rrUp = list(filter(lambda x: x[4] == 'Up' and x[2] == 'RR', values))
            rrDown = list(filter(lambda x: x[4] == 'Down' and x[2] == 'RR', values))
            #srrUp = list(filter(lambda x: x[4] == 'Up' and x[2] == 'Spinning RR', values))
            #srrDown = list(filter(lambda x: x[4] == 'Down' and x[2] == 'Spinning RR', values))

            row = {
                'DispatchDay': k['dispatchDay'],
                'DispatchPeriod': k['dispatchPeriod'],
                'EntityName': k['entityName'],
                'Version': version
            }

            if len(fcrUp) > 0:
                row['ValueFcrup'] = fcrUp[0][5]
            if len(fcrDown) > 0:
                row['ValueFcrdown'] = fcrDown[0][5]
            if len(afrrUp) > 0:
                row['ValueAFrrup'] = afrrUp[0][5]
            if len(afrrDown) > 0:
                row['ValueAFrrdown'] = afrrDown[0][5]
            if len(mfrrUp) > 0:
                row['ValueMFrrup'] = mfrrUp[0][5]
            if len(mfrrDown) > 0:
                row['ValueMFrrdown'] = mfrrDown[0][5]
            if len(rrUp) > 0:
                row['ValueRrup'] = rrUp[0][5]
            if len(rrDown) > 0:
                row['ValueRrdown'] = rrDown[0][5]
            # if len(srrUp) > 0:
            #     row['ValueSRrup'] = srrUp[0][5]
            # if len(srrDown) > 0:
            #     row['ValueSRrdown'] = srrDown[0][5]

            results.append(row)

        results = fill_dispatch_periods_for_isp_market(file.TargetDateFrom, results)

        return envelope.getSuccessEnvelope(getParsingMetadataPayload(results, getMetadata(True, file.Id)))
    except Exception as e:
        message = f'Failed to parse {path}: {e}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

def getISPScheduleFromISPResults(dateFrom: str, dateTo : str, isp: int = None, adhoc: bool =None):
    files = getISPResultsFiles(dateFrom, dateTo,isp, adhoc)

    for file in files:
        yield getIspScheduleActualFromFile(file)
    
def getIspScheduleActualFromFile(file: CustomFile):
    results = []
    version = file.PublicationDate

    path = getFileNameFromInfo(file)
    if not os.path.exists(path):
        message = f'{path} does not exist'
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        ISP_DataFrame = pd.read_excel(path, skiprows = 1, sheet_name=0)
        dfStyle = load_workbook(path, data_only = True)
    except Exception as e:
        message = f'Failed to read {path}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)
    
    try:
        first_column = ISP_DataFrame.columns[0]

        ISPSchedules = []

        index_thermal_units = ISP_DataFrame.index[ISP_DataFrame[first_column] == 'Thermal Units'].values[-1]
        getISPResultsDataFromRange(ISP_DataFrame, index_thermal_units + 1, len(ISP_DataFrame.values), "Thermal", ISPSchedules, 1, dfStyle=dfStyle)

        index_hydro_units = ISP_DataFrame.index[ISP_DataFrame[first_column] == 'Hydro Units'].values[-1]
        getISPResultsDataFromRange(ISP_DataFrame, index_hydro_units + 1, len(ISP_DataFrame.values), "Hydro", ISPSchedules, 1, dfStyle=dfStyle)

        index_res_units = ISP_DataFrame.index[ISP_DataFrame[first_column] == 'RES Units'].values[-1]
        getISPResultsDataFromRange(ISP_DataFrame, index_res_units + 1, len(ISP_DataFrame.values), "RES", ISPSchedules, 1, dfStyle=dfStyle)

        index_dispatchable_load = ISP_DataFrame.index[ISP_DataFrame[first_column] == 'Dispatchable Load'].values[-1]
        getISPResultsDataFromRange(ISP_DataFrame, index_dispatchable_load + 1, len(ISP_DataFrame.values), "Pumping", ISPSchedules, 1, dfStyle=dfStyle)
        
        demand_response = ISP_DataFrame.index[ISP_DataFrame[first_column] == 'Demand Response']
        if len(demand_response.values) > 0:
            index_demand_response = demand_response.values[0]
            getISPResultsDataFromRange(ISP_DataFrame, index_demand_response + 1, len(ISP_DataFrame.values), "Demand Response", ISPSchedules, 1, dfStyle=dfStyle)

        for row in ISPSchedules:
            results.append({
                'DispatchDay': datetime.strptime(row[0], '%Y-%m-%d'),
                'DispatchPeriod': row[1],
                'EntityName': row[3],
                'Value': row[4],
                'EntityState': row[5],
                'Version': version,
                #'VersionDescription': versionDescription
            })
        
        results = fill_dispatch_periods_for_isp_market(file.TargetDateFrom, results)

        return envelope.getSuccessEnvelope(getParsingMetadataPayload(results, getMetadata(True, file.Id)))
    except Exception as e:
        message = f'Failed to parse {path}: {e}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

def getIspSystemLoadFromISPResults(dateFrom: str, dateTo : str, isp: int = None, adhoc: bool =None):
    files = getISPResultsFiles(dateFrom, dateTo,isp, adhoc)

    for file in files:
        yield getIspSystemLoadFromFile(file)
    
def getIspSystemLoadFromFile(file: CustomFile):
    results = []
    version = file.PublicationDate

    path = getFileNameFromInfo(file)
    if not os.path.exists(path):
        message = f'{path} does not exist'
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        ISP_DataFrame = pd.read_excel(path, skiprows = 1, sheet_name=0)
    except Exception as e:
        message = f'Failed to read {path}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)
    
    try:
        first_column = ISP_DataFrame.columns[0]

        index_system_load = ISP_DataFrame.index[ISP_DataFrame[first_column] == 'System Load+Losses'].values[-1]

        ISPSystemLoad = []
        
        getISPResultsDataFromRange(ISP_DataFrame, index_system_load, index_system_load + 1, "Thermal", ISPSystemLoad)

        for row in ISPSystemLoad:

            results.append({
                'DispatchDay': datetime.strptime(row[0], '%Y-%m-%d'),
                'DispatchPeriod': row[1],
                'ZoneId': 1,
                'Value': row[4],
                'Version': version
            })

        results = fill_dispatch_periods_for_isp_market(file.TargetDateFrom, results)

        return envelope.getSuccessEnvelope(getParsingMetadataPayload(results, getMetadata(True, file.Id)))
    except Exception as e:
        message = f'Failed to parse {path}: {e}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

def getIspVirtualScheduleConfiguration(dateFrom: str, dateTo : str, isp: int = None, adhoc: bool =None):
    files = getISPResultsFiles(dateFrom, dateTo, isp, adhoc)
    
    for file in files:
        yield getIspVirtualScheduleConfigurationFromFile(file)

def getIspVirtualScheduleConfigurationFromFile(file: CustomFile):
    results = []
    version = file.PublicationDate

    path = getFileNameFromInfo(file)
    if not os.path.exists(path):
        message = f'{path} does not exist'
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        ISP_DataFrame = pd.read_excel(path, sheet_name=2)
    except Exception as e:
        message = f'Failed to read {path}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)
    
    try:
        ISPSchedules = []
        getCCGTFromRange(ISP_DataFrame, 1, ISP_DataFrame.shape[0], ISPSchedules)

        for row in ISPSchedules:
            results.append({
                'DispatchDay': row[0],
                'DispatchPeriod': row[1],
                'PrimaryEntityName': row[2],
                'VirtualEntityName': row[3],
                'Version': version,
                #'VersionDescription': versionDescription
            })
        
        results = fill_dispatch_periods_for_isp_market(row[0], results)

        return envelope.getSuccessEnvelope(getParsingMetadataPayload(results, getMetadata(True, file.Id)))
    except Exception as e:
        message = f'Failed to parse {path}: {e}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

def getISPGenericConstraints(dateFrom: str, dateTo : str, isp: int = None, adhoc: bool =None):
    files = getISPResultsFiles(dateFrom, dateTo, isp, adhoc)
    
    for file in files:
        yield getISPGenericConstraintsFromFile(file)

def getISPGenericConstraintsFromFile(file: CustomFile):
    results = []
    version = file.PublicationDate

    def addVersionInfo(data: dict):
        data['Version'] = version
        #data['VersionDescription'] = versionDescription
        data['DispatchDay'] = data['StartTime'].date()

        if math.isnan(data['Reason']):
            data['Reason'] = None
        return data
    
    if version > datetime(2020, 12, 1, 15, 45):
        columns = ['ConstraintId', 'StartTime', 'EndTime', 'ConstraintType', 'Limit', 'EntityName', 'EnergyFactor', 'FcrUpFactor', 'FcrDownFactor', 'AfrrUpFactor','AfrrDownFactor','MfrrUpFactor','MfrrDownFactor','RrUpFactor','RrDownFactor','Reason']
    else:
        columns = ['ConstraintId', 'StartTime', 'EndTime', 'ConstraintType', 'Limit', 'EntityName', 'EnergyFactor', 'RrUpFactor', 'Reason']

    path = getFileNameFromInfo(file)
    if not os.path.exists(path):
        message = f'{path} does not exist'
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        ISP_DataFrame = pd.read_excel(path, skiprows = 1, sheet_name=7, names = columns)
    except Exception as e:
        message = f'Failed to read {path}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)
    
    try:
        data = list(map(addVersionInfo, ISP_DataFrame.to_dict('records')))

        results.extend(data)

        return envelope.getSuccessEnvelope(getParsingMetadataPayload(results, getMetadata(True, file.Id)))
    except Exception as e:
        message = f'Failed to parse {path}: {e}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

def getISPBalancingEnergyPrices(dateFrom: str, dateTo : str, isp: int = None, adhoc: bool =None):
    files = getISPResultsFiles(dateFrom, dateTo, isp, adhoc)
    
    for file in files:
        yield getISPBalancingEnergyPricesFromFile(file)

def getISPBalancingEnergyPricesFromFile(file: CustomFile):
    results = []
    version = file.PublicationDate

    path = getFileNameFromInfo(file)
    if not os.path.exists(path):
        message = f'{path} does not exist'
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        ISP_DataFrame = pd.read_excel(path, skiprows = 1, sheet_name=0)
        dfStyle = load_workbook(path, data_only = True)
    except Exception as e:
        message = f'Failed to read {path}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)
    
    try:
        first_column = ISP_DataFrame.columns[0]
        first_date = ISP_DataFrame.columns[1]

        index_price_energy_up = ISP_DataFrame.index[ISP_DataFrame[first_column] == 'Price Energy Up'].values[0]
        index_price_energy_down = ISP_DataFrame.index[ISP_DataFrame[first_column] == 'Price Energy Down'].values[0]

        first_date_pd = pd.Timestamp(first_date)
        period = getTimeDifferncePandas(first_date_pd, seconds_granularity= 900)
        #period = getDispatchPeriodFromHourString(first_date.strftime('%H:%M'))

        for i in range(1, len(ISP_DataFrame.columns) - 1):
            up_data = ISP_DataFrame.iloc[index_price_energy_up, :].values
            down_data = ISP_DataFrame.iloc[index_price_energy_down, :].values

            dispatchDate = first_date.date()
            if type(ISP_DataFrame.columns[i]) is not datetime:
                if not checkDateToString(ISP_DataFrame.columns[i]):
                    continue

            
            #for j in range(1,3):
            results.append({
                'DispatchDay': dispatchDate,
                'DispatchPeriod': period,#2*(period-1) + j,
                'Version': version,
                #'VersionDescription': versionDescription,
                'ZoneId': 1,
                'PriceUp': up_data[i],
                'PriceDown': down_data[i]
            })
            period += 1
            
        results = fill_dispatch_periods_for_isp_market(dispatchDate, results)

        return envelope.getSuccessEnvelope(getParsingMetadataPayload(results, getMetadata(True, file.Id)))
    except Exception as e:
        message = f'Failed to parse {path}: {e}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

def getISPSystemImbalance(dateFrom: str, dateTo : str, isp: int = None, adhoc: bool =None):
    files = getISPResultsFiles(dateFrom, dateTo, isp, adhoc)
    
    for file in files:
        yield getISPSystemImbalanceFromFile(file)

def getISPSystemImbalanceFromFile(file: CustomFile):
    results = []
    version = file.PublicationDate

    path = getFileNameFromInfo(file)
    if not os.path.exists(path):
        message = f'{path} does not exist'
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        ISP_DataFrame = pd.read_excel(path, skiprows = 1, sheet_name=0)
        dfStyle = load_workbook(path, data_only = True)
    except Exception as e:
        message = f'Failed to read {path}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)
    
    try:
        first_column = ISP_DataFrame.columns[0]
        first_date = ISP_DataFrame.columns[1]

        index_energy_up = ISP_DataFrame.index[ISP_DataFrame[first_column] == 'Cleared Energy Up'].values[0]
        index_energy_down = ISP_DataFrame.index[ISP_DataFrame[first_column] == 'Cleared Energy Down'].values[0]

        first_date_pd = pd.Timestamp(first_date)
        period = getTimeDifferncePandas(first_date_pd, seconds_granularity= 900)
        #period = getDispatchPeriodFromHourString(first_date.strftime('%H:%M'))
        for i in range(1, len(ISP_DataFrame.columns) - 1):
            dispatchDate = first_date.date()

            up_data = ISP_DataFrame.iloc[index_energy_up, :].values
            down_data = ISP_DataFrame.iloc[index_energy_down, :].values

            if type(ISP_DataFrame.columns[i]) is not datetime:
                if not checkDateToString(ISP_DataFrame.columns[i]):
                    continue
        
            
            results.append({
                'DispatchDay': dispatchDate,
                'DispatchPeriod': period,
                'Version': version,
                #'VersionDescription': versionDescription,
                'ZoneId': 1,
                'ClearedEnergyUp': up_data[i],
                'ClearedEnergyDown': down_data[i]
            })
            period += 1

        results = fill_dispatch_periods_for_isp_market(dispatchDate, results)

        return envelope.getSuccessEnvelope(getParsingMetadataPayload(results, getMetadata(True, file.Id)))
    except Exception as e:
        message = f'Failed to parse {path}: {e}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

def getISPReservePrices(dateFrom: str, dateTo : str, isp: int = None, adhoc: bool =None):
    files = getISPResultsFiles(dateFrom, dateTo, isp, adhoc)
    
    for file in files:
        yield getISPReservePricesFromFile(file)

def getISPReservePricesFromFile(file: CustomFile):
    results = []
    PricesData = []
    version = file.PublicationDate

    path = getFileNameFromInfo(file)
    if not os.path.exists(path):
        message = f'{path} does not exist'
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        xls = pd.ExcelFile(path)
    except Exception as e:
        message = f'Failed to read {path}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)
    
    try:
        def parseSheet(reserve: str, sheet: int):
            ISPResults = pd.read_excel(xls, skiprows = [0], sheet_name=sheet)
        
            first_column = ISPResults.columns[0]

            index_price_up = ISPResults.index[ISPResults[first_column] == ('Price Up' if reserve != 'RR' else 'RR Price Up')].values[0]
            index_price_down = ISPResults.index[ISPResults[first_column] == ('Price Down' if reserve != 'RR' else 'RR Price Down')].values[0]

            first_date = ISPResults.columns[1]
            first_date_pd = pd.Timestamp(first_date)
            period = getTimeDifferncePandas(first_date_pd, seconds_granularity= 900)
            #dispatchPeriod = getDispatchPeriodFromHourString(first_date.strftime('%H:%M'))
            for i in range(1, len(ISPResults.columns)):
                if type(ISPResults.columns[i]) is not datetime:
                    if not checkDateToString(ISPResults.columns[i]):
                        continue
                

                up_data = ISPResults.iloc[index_price_up, :].values
                down_data = ISPResults.iloc[index_price_down, :].values

                PricesData.append({
                    'DispatchDay': first_date.date(),
                    'DispatchPeriod': period,
                    'Version': version,
                    #'VersionDescription': versionDescription,
                    'ReserveType': reserve,
                    'PriceUp': up_data[i],
                    'PriceDown': down_data[i]
                })
                period += 1

        parseSheet('FCR', 3)
        parseSheet('aFRR', 4)
        parseSheet('mFRR', 5)
        parseSheet('RR', 6)

        key = lambda x: {
            'dispatchDay': x['DispatchDay'],
            'dispatchPeriod': x['DispatchPeriod']
        }

        PricesData.sort(key=lambda x: x['DispatchPeriod'])
        PricesData.sort(key=lambda x: x['DispatchDay'])
        
        for k,v in groupby(PricesData,key=key):
            values = list(v)

            fcr = list(filter(lambda x: x['ReserveType'] == 'FCR', values))
            afrr = list(filter(lambda x: x['ReserveType'] == 'aFRR', values))
            mfrr = list(filter(lambda x: x['ReserveType'] == 'mFRR', values))
            rr = list(filter(lambda x: x['ReserveType'] == 'RR', values))
            #srr = list(filter(lambda x: x['ReserveType'] == 'Spinning RR', values))

            
            row = {
                'DispatchDay': k['dispatchDay'],
                'DispatchPeriod': (k['dispatchPeriod']) ,
                'ZoneId': 1,
                'Version': version,
                #'VersionDescription': versionDescription
            }

            if len(fcr) > 0:
                row['PriceFcrUp'] = fcr[0]['PriceUp']
                row['PriceFcrDown'] = fcr[0]['PriceDown']
            if len(afrr) > 0:
                row['PriceAfrrUp'] = afrr[0]['PriceUp']
                row['PriceAfrrDown'] = afrr[0]['PriceDown']
            if len(mfrr) > 0:
                row['PriceMfrrUp'] = mfrr[0]['PriceUp']
                row['PriceMfrrDown'] = mfrr[0]['PriceDown']
            if len(rr) > 0:
                row['PriceRrUp'] = rr[0]['PriceUp']
                row['PriceRrDown'] = rr[0]['PriceDown']

            results.append(row)

        results = fill_dispatch_periods_for_isp_market(file.TargetDateFrom, results)

        return envelope.getSuccessEnvelope(getParsingMetadataPayload(results, getMetadata(True, file.Id)))
    except Exception as e:
        message = f'Failed to parse {path}: {e}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)
##


## ISP Forecasts
def getIspLoadForecasts(dateFrom: datetime, dateTo: datetime, isp: int = None, version: int =None):
    files = getISPFilesWithPrefix("DayAheadLoadForecast", dateFrom, dateTo, isp, version)
    
    for file in files:
        yield getIspLoadForecastsFromFile(file)

def getIspLoadForecastsFromFile(file: CustomFile):
    results = []
    version = file.PublicationDate

    path = getFileNameFromInfo(file)
    if not os.path.exists(path):
        message = f'{path} does not exist'
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        market_change_flag = file.TargetDateFrom >= market_change_date

        df = pd.read_excel(path, skiprows=2 if not market_change_flag else None)
        date_df = pd.read_excel(path)
        excel_date = getDateFromDf(date_df, market_change_flag)
    except Exception as e:
        message = f'Failed to read {path}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)
    
    try:
        dispatchPeriod = getDispatchPeriodFromHourString(file.TargetDateFrom.strftime('%H:%M'))
        for t in df.columns[1:]:
            if t.startswith('Unnamed'):
                continue

            value = df[str(t)][0]
            results.append({'dispatchDay': excel_date,
                'dispatchPeriod': dispatchPeriod,
                'version': version,
                #'versionDescription': versionDescription,
                'zoneId': 1,
                'fileId': file.Id,
                'loadForecast': float(value)
            })
            dispatchPeriod += 1

        results = fill_dispatch_periods_for_isp_market(file.TargetDateFrom, results)

        return envelope.getSuccessEnvelope(getParsingMetadataPayload(results, getMetadata(True, file.Id)))
    except Exception as e:
        message = f'Failed to parse {path}: {e}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

def getIspResForecasts(dateFrom: datetime, dateTo: datetime, isp: int = None, version: int =None):
    files = getISPFilesWithPrefix("DayAheadRESForecast", dateFrom, dateTo, isp, version)
    
    for file in files:
        yield getIspResForecastsFromFile(file)

def getIspResForecastsFromFile(file: CustomFile):
    results = []
    version = file.PublicationDate

    path = getFileNameFromInfo(file)
    if not os.path.exists(path):
        message = f'{path} does not exist'
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        market_change_flag = file.TargetDateFrom >= market_change_date

        df = pd.read_excel(path, skiprows=2 if not market_change_flag else None)
        date_df = pd.read_excel(path)
        excel_date = getDateFromDf(date_df, market_change_flag)
    except Exception as e:
        message = f'Failed to read {path}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)
    
    try:
        
        dispatchPeriod = getDispatchPeriodFromHourString(excel_date.strftime('%H:%M'))
        for t in df.columns:
            if t.startswith('Unnamed'):
                continue

            value = df[str(t)][0]
            results.append({'dispatchDay': excel_date.strftime('%Y-%m-%d'),
                'dispatchPeriod': dispatchPeriod,
                'version': version,
                #'versionDescription': versionDescription,
                'zoneId': 1,
                'fileId': file.Id,
                'resForecast': float(value)
            })
            dispatchPeriod += 1

        results = fill_dispatch_periods_for_isp_market(excel_date, results)

        return envelope.getSuccessEnvelope(getParsingMetadataPayload(results, getMetadata(True, file.Id)))
    except Exception as e:
        message = f'Failed to parse {path}: {e}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

def getIspWeekAheadLoadForecasts(dateFrom: datetime, dateTo: datetime):
    files = getFileUrlsFromApi("ISPWeekAheadLoadForecast", dateFrom, dateTo) 

    for file in files:
        yield getIspWeekAheadLoadForecastsFromFile(file)

def getIspWeekAheadLoadForecastsFromFile(file: CustomFile):
    results = []
    version = file.PublicationDate

    path = getFileNameFromInfo(file)
    if not os.path.exists(path):
        message = f'{path} does not exist'
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        df = pd.read_excel(path, skiprows=[0])
    except Exception as e:
        message = f'Failed to read {path}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)
    
    try:
        results = []
        for i in range(df.shape[0]):
            date = df.iloc[i,2]
            day_results = []
            for t in df.columns:
                if t.startswith('Unnamed'):
                    continue
                
                value = df[str(t)][i]
                if value == 0 or pd.isna(value):
                    continue

                day_results.append({
                    'dispatchDay': date,
                    'dispatchPeriod': int(t),
                    'versionDescription': '',
                    'version': version,
                    'zoneId': 1,
                    'fileId': file.Id,
                    'loadForecast': float(value)
                })

            results.extend(day_results)

        return envelope.getSuccessEnvelope(getParsingMetadataPayload(results, getMetadata(True, file.Id)))
    except Exception as e:
        message = f'Failed to parse {path}: {e}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)
##


## Balancing Settlements - IMBABE

def getBalancingEnergySettlements(dateFrom: datetime, dateTo: datetime):
    files = getFileUrlsFromApi("IMBABE", dateFrom, dateTo) 

    for file in files:
        yield getBalancingEnergySettlementsFromFile(file)

def getBalancingEnergySettlementsFromFile(file: CustomFile):
    results = []
    
    version = file.PublicationDate

    path = getFileNameFromInfo(file)
    if not os.path.exists(path):
        message = f'{path} does not exist'
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        df = pd.read_excel(path, header=0)

        # aliviate unwanted text data in excel file by keeping rows with datetime.        
        # Drop completely empty rows
        df = df.dropna(how="all")
        mask = pd.to_datetime(df.iloc[:,0], errors="coerce").notna()

        df = df.loc[mask[mask==True].index]
    except Exception as e:
        message = f'Failed to read {path}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        timestamps = pd.read_excel(path).loc[mask[mask==True].index]['STARTDATE']
        time_to_pd = pd.to_datetime(timestamps)
        time_infered = time_to_pd.dt.tz_localize('Europe/Berlin',ambiguous="infer")
        for i,(_,row) in enumerate(df.iterrows()):
            dateStart = row["STARTDATE"]
            
            if isinstance(dateStart, str):
                dateStart = datetime.strptime(dateStart, '%d/%m/%Y %H:%M')

            hour_dispatch = getTimeDifferncePandasBallancingCap(time_infered[i])
            minute_dispatch = int(dateStart.minute) // 15

            BalancingEnergyIDEV = float(row['Balancing Energy IDEV (MWh)']) if 'Balancing Energy IDEV (MWh)' in row and not math.isnan(row['Balancing Energy IDEV (MWh)']) else 0
            BalancingEnergyUDEV = float(row['Balancing Energy UDEV (MWh)']) if 'Balancing Energy UDEV (MWh)' in row and not math.isnan(row['Balancing Energy UDEV (MWh)']) else 0
            
            if pd.isna(row['mFRR Price Down (€/MWh)']):
                row['mFRR Price Down (€/MWh)'] = 0


            # alleviate unwanted spaces in column names because admie sometimes adds whitespaces without warning.
            row.index = row.index.str.replace(r"\s+", "", regex=True)
            
            # Parse new optional columns
            ISPEnergyDPkDF = None
            if '(ΔP+kΔf)"ISPenergy"(MWh)' in row and not pd.isna(row['(ΔP+kΔf)"ISPenergy"(MWh)']):
                ISPEnergyDPkDF = float(row['(ΔP+kΔf)"ISPenergy"(MWh)'])
            
            SystemDeviation = None
            if 'Systemdeviation(MWh)' in row and not pd.isna(row['Systemdeviation(MWh)']):
                SystemDeviation = float(row['Systemdeviation(MWh)'])
            
            CurtFlag = None
            if 'Curt_Flag' in row and not pd.isna(row['Curt_Flag']):
                CurtFlag = bool(int(row['Curt_Flag'])) 
            results.append(BalancingEnergyData(**{
                "DispatchDay": dateStart.date(),
                "DispatchPeriod": 4*(hour_dispatch-1) + minute_dispatch + 1,
                "Version": version,
                "ZoneId": 1,
                "totalActivatedBalancingEnergyUp": row["TotalActivatedBalancingEnergyUP(MWh)"],
                "totalActivatedBalancingEnergyDown":row["TotalActivatedBalancingEnergyDown(MWh)"],
                "imbalancePrice": row["ImbalancePrice(€/MWh)"],
                "mFrrUpPrice": row["mFRRPriceUP(€/MWh)"],
                "mFrrDownPrice": row["mFRRPriceDown(€/MWh)"],
                "upliftAccount1": row["UpliftAccount1(€/MWh)"],
                "upliftAccount2": row["UpliftAccount2(€/MWh)"],
                "upliftAccount3": row["UpliftAccount3(€/MWh)"],
                "balancingEnergyIdev": BalancingEnergyIDEV,
                "balancingEnergyUdev": BalancingEnergyUDEV,
                "ISPEnergyDPkDF": ISPEnergyDPkDF,
                "SystemDeviation": SystemDeviation,
                "CurtFlag": CurtFlag,
                'FileId': file.Id
            }))

        return envelope.getSuccessEnvelope(getParsingMetadataPayload(results, getMetadata(True, file.Id)))
    except Exception as e:
        message = f'Failed to parse {path}: {e}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)
##


## SCADA Documents Parsing

def getImportExportFromSCADA(dateFrom: datetime, dateTo: datetime):
    files = getFileUrlsFromApi("SystemRealizationSCADA", dateFrom, dateTo) 

    for file in files:
        yield getImportExportFromSCADAFromFile(file)

def getImportExportFromSCADAFromFile(file: CustomFile):
    results = []

    path = getFileNameFromInfo(file)
    if not os.path.exists(path):
        message = f'{path} does not exist'
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        SystemRealizationSCADA = ReadSCADAFile(path)
    except Exception as e:
        message = f'Failed to read {path}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        Date = file.TargetDateFrom
        names = getSCADACountryNames(Date)
        
        date_to_pd = pd.Timestamp(Date)
        hours = getTimeDifferncePandas(date_to_pd - timedelta(hours=1) + timedelta(days=1))

        dispatchDate = EETtoCET(Date)
        dispatch_columns = SystemRealizationSCADA.iloc[1,2:2 + hours]
        for name in names:
            
            Imp = SystemRealizationSCADA[SystemRealizationSCADA.iloc[:, 1] == name + "_IMP"].values[0][2:2+hours]
            Exp = SystemRealizationSCADA[SystemRealizationSCADA.iloc[:, 1] == name + "_EXP"].values[0][2:2+hours]

            offset = 0
            
            for i in range(len(dispatch_columns)):
                
                hour_column = dispatch_columns[i]
                #check if we can parse it to int from string
                if not isParsableToInt(hour_column):
                        continue 
                
                try:
                    dispatchDate = date_to_pd - timedelta(hours=2) + timedelta(hours=int(hour_column) + offset)
                    dispatch_hours = getTimeDifferncePandas(dispatchDate)
                except pytz.exceptions.NonExistentTimeError:
                    #catch error in order to handle scada summer dst issues. add +1 offset
                    offset = 1
                    dispatchDate = date_to_pd - timedelta(hours=2) + timedelta(hours=int(hour_column) + offset)
                    dispatch_hours = getTimeDifferncePandas(dispatchDate)
                
                for j in range(1, 5):
                    results.append({
                        "DispatchDay": dispatchDate.date(),
                        "DispatchPeriod": 4*(dispatch_hours - 1) + j,
                        "ImportValue": float(Imp[i]),
                        "ExportValue": float(Exp[i]),
                        "CountryName": name,
                        'FileId': file.Id,
                        "Version":  file.PublicationDate
                    })
                    #hour_column += 1

        return envelope.getSuccessEnvelope(getParsingMetadataPayload(results, getMetadata(True, file.Id)))
    except Exception as e:
        message = f'Failed to parse {path}: {e}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)
    


def getSCADAAggregatedFromFile(file) -> Envelope[List[VirtualFileMetadataPayload[List[ScadaAggregatedProdModel]]]]:
    results = []

    path = getFileNameFromInfo(file) 
    if not os.path.exists(path):
        message = f'{path} does not exist'
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        SystemRealizationSCADA = ReadSCADAFile(path)
    except Exception as e:
        message = f'Failed to read {path}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        Date = file.TargetDateFrom
        names = getSCADAAggregatedNames()
        
        date_to_pd = pd.Timestamp(Date)
        hours = getTimeDifferncePandas(date_to_pd - timedelta(hours=1) + timedelta(days=1))

        # do Date astimezone to Cet and convert to UTC and Convert the localized date to UTC
        et_timezone = pytz.timezone('Europe/Athens')
        localized_date = et_timezone.localize(Date)
        dispatchDate = localized_date.astimezone(pytz.utc)
        dispatch_columns = SystemRealizationSCADA.iloc[1,2:2 + hours]
        
        #keep rows of data in df only in second column contains variable names
        aggregated_data_df = SystemRealizationSCADA[SystemRealizationSCADA.iloc[:, 1].isin(names)] 

        for i in range(len(dispatch_columns)):
            data_dict = {
                "Timestamp": (dispatchDate + timedelta(hours = i)).strftime('%Y-%m-%dT%H:%M:%SZ'),
                "NetLoad": float(aggregated_data_df.loc[aggregated_data_df.iloc[:,1]=="NET LOAD"].iloc[:,2+i]),
                "NetLoadWithoutCrete": float(aggregated_data_df.loc[aggregated_data_df.iloc[:,1]=="NET LOAD WITHOUT CRETE"].iloc[:,2+i]),
                "TotalLignite": float(aggregated_data_df.loc[aggregated_data_df.iloc[:,1]=="TOTAL LIGNITE"].iloc[:,2+i]),
                "TotalOil": float(aggregated_data_df.loc[aggregated_data_df.iloc[:,1]=="ΣΥΝΟΛΟ ΠΕΤΡΕΛΑΙΚΩΝ"].iloc[:,2+i]),
                "TotalGas": float(aggregated_data_df.loc[aggregated_data_df.iloc[:,1]=="TOTAL GAS"].iloc[:,2+i]),
                "TotalHydro": float(aggregated_data_df.loc[aggregated_data_df.iloc[:,1]=="TOTAL HYDRO"].iloc[:,2+i]),
                "TotalRes": float(aggregated_data_df.loc[aggregated_data_df.iloc[:,1]=="TOTAL RES"].iloc[:,2+i]),
                'FileId': file.Id,
                "Version":  file.PublicationDate
            }
            data = ScadaAggregatedProdModel(**data_dict)
            results.append(data)
        return envelope.getSuccessEnvelope(getParsingMetadataPayload(results, getMetadata(True, file.Id)))
    except Exception as e:
        message = f'Failed to parse {path}: {e}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)



def getScadaResProduction(dateFrom: str, dateTo : str):
    files = getFileUrlsFromApi("SystemRealizationSCADA", dateFrom, dateTo) 

    for file in files:
        yield getScadaResProductionFromFile(file)

def getScadaResProductionFromFile(file: CustomFile):
    results = []

    path = getFileNameFromInfo(file)
    if not os.path.exists(path):
        message = f'{path} does not exist'
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        SystemRealizationSCADA = ReadSCADAFile(path)
    except Exception as e:
        message = f'Failed to read {path}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        Date = file.TargetDateFrom

        res_begin = SystemRealizationSCADA[SystemRealizationSCADA.iloc[:, 1].str.contains("ΑΝΑΝΕΩΣΙΜΑ") == True].index[0] + 1
        res_end = SystemRealizationSCADA[(SystemRealizationSCADA.iloc[:, 1].str.contains("TOTAL RES") == True) | (SystemRealizationSCADA.iloc[:, 1].str.contains("ΣΥΝΟΛΟ ΑΙΟΛΙΚΩΝ") == True)].index[0] - 1
        ResProduction = []
        getScadaDataFromRange(SystemRealizationSCADA, res_begin, res_end, Date, ResProduction)

        for row in ResProduction:
            for period in range(1, 5):
                results.append({
                    'DispatchDay': row[0],
                    'DispatchPeriod': (int(row[1])-1) * 4 + period,
                    'Value': float(row[3]),
                    'Version':file.PublicationDate,
                    'FileId': file.Id,
                    'EntityName': row[2],
                })

        return envelope.getSuccessEnvelope(getParsingMetadataPayload(results, getMetadata(True, file.Id)))
    except Exception as e:
        message = f'Failed to parse {path}: {e}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)


def getScadaEntityProduction(dateFrom: str, dateTo : str):
    files = getFileUrlsFromApi("SystemRealizationSCADA", dateFrom, dateTo) 

    for file in files:
        yield getScadaEntityProductionFromFile(file)

def getScadaEntityProductionFromFile(file: CustomFile):
    results = []

    path = getFileNameFromInfo(file)
    if not os.path.exists(path):
        message = f'{path} does not exist'
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        SystemRealizationSCADA = ReadSCADAFile(path)
    except Exception as e:
        message = f'Failed to read {path}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        Date = file.TargetDateFrom

        getOilProductionFromScada(file, results, SystemRealizationSCADA, Date)
        getScadaHydroData(file, results, SystemRealizationSCADA, Date)
        getNaturalGasFromScada(file, results, SystemRealizationSCADA, Date)
        getThermalProductionFromScada(file, results, SystemRealizationSCADA, Date)

        return envelope.getSuccessEnvelope(getParsingMetadataPayload(results, getMetadata(True, file.Id)))
    except Exception as e:
        message = f'Failed to parse {path}: {e}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

def getScadaHydroProduction(dateFrom: str, dateTo : str):
    files = getFileUrlsFromApi("SystemRealizationSCADA", dateFrom, dateTo) 

    for file in files:
        yield getScadaHydroProductionFromFile(file)

def getScadaHydroProductionFromFile(file: CustomFile):
    results = []

    path = getFileNameFromInfo(file)
    if not os.path.exists(path):
        message = f'{path} does not exist'
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        SystemRealizationSCADA = ReadSCADAFile(path)
    except Exception as e:
        message = f'Failed to read {path}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        Date = file.TargetDateFrom
        
        getScadaHydroData(file, results, SystemRealizationSCADA, Date)
        
        return envelope.getSuccessEnvelope(getParsingMetadataPayload(results, getMetadata(True, file.Id)))
    except Exception as e:
        message = f'Failed to parse {path}: {e}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

def getScadaNaturalGasProduction(dateFrom: str, dateTo : str):
    files = getFileUrlsFromApi("SystemRealizationSCADA", dateFrom, dateTo) 

    for file in files:
        yield getScadaNaturalGasProductionFromFile(file)
    
def getScadaNaturalGasProductionFromFile(file: CustomFile):
    results = []

    path = getFileNameFromInfo(file)
    if not os.path.exists(path):
        message = f'{path} does not exist'
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        SystemRealizationSCADA = ReadSCADAFile(path)
    except Exception as e:
        message = f'Failed to read {path}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        Date = file.TargetDateFrom
        getNaturalGasFromScada(file, results, SystemRealizationSCADA, Date)

        return envelope.getSuccessEnvelope(getParsingMetadataPayload(results, getMetadata(True, file.Id)))
    except Exception as e:
        message = f'Failed to parse {path}: {e}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

def getScadaThermoProduction(dateFrom: str, dateTo : str):
    files = getFileUrlsFromApi("SystemRealizationSCADA", dateFrom, dateTo) 

    for file in files:
        yield getScadaThermoProductionFromFile(file)

def getScadaThermoProductionFromFile(file: CustomFile):
    results = []

    path = getFileNameFromInfo(file)
    if not os.path.exists(path):
        message = f'{path} does not exist'
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        SystemRealizationSCADA = ReadSCADAFile(path)
    except Exception as e:
        message = f'Failed to read {path}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        Date = file.TargetDateFrom

        getThermalProductionFromScada(file, results, SystemRealizationSCADA, Date)
    
        return envelope.getSuccessEnvelope(getParsingMetadataPayload(results, getMetadata(True, file.Id)))
    except Exception as e:
        message = f'Failed to parse {path}: {e}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

def getScadaHVLoad(dateFrom: str, dateTo : str):
    files = getFileUrlsFromApi("SystemRealizationSCADA", dateFrom, dateTo) 

    for file in files:
        yield getScadaHVLoadFromFile(file)

def getScadaHVLoadFromFile(file: CustomFile):
    results = []

    path = getFileNameFromInfo(file)
    if not os.path.exists(path):
        message = f'{path} does not exist'
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        SystemRealizationSCADA = ReadSCADAFile(path)
    except Exception as e:
        message = f'Failed to read {path}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        Date = file.TargetDateFrom
        HV_index = SystemRealizationSCADA[SystemRealizationSCADA.iloc[:, 1].str.contains("ΚΑΤΑΝΑΛΩΤΕΣ 150") == True].index[0]
        HVLoad = []
        getScadaDataFromRange(SystemRealizationSCADA, HV_index, HV_index, Date, HVLoad)

        for row in HVLoad:
            for period in range(1, 5): 
                results.append({
                        'DispatchDay': row[0].strftime('%Y-%m-%d'),
                        'DispatchPeriod': (int(row[1]) - 1) * 4 + period,
                        'FileId': file.Id,
                        'ZoneId': 1,
                        'Value': float(row[3]) if not math.isnan(row[3]) else 0,
                        'Version': file.PublicationDate
                    })
        return envelope.getSuccessEnvelope(getParsingMetadataPayload(results, getMetadata(True, file.Id)))
    except Exception as e:
        message = f'Failed to parse {path}: {e}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)


def getScadaSystemLoadRealization(dateFrom: str, dateTo : str):
    files = getFileUrlsFromApi("SystemRealizationSCADA", dateFrom, dateTo) 

    for file in files:
        yield getScadaSystemLoadRealizationFromFile(file)

def getScadaSystemLoadRealizationFromFile(file: CustomFile):
    results = []

    path = getFileNameFromInfo(file)
    if not os.path.exists(path):
        message = f'{path} does not exist'
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        SystemRealizationSCADA = ReadSCADAFile(path)
    except Exception as e:
        message = f'Failed to read {path}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        Date = file.TargetDateFrom
        Gross_index = SystemRealizationSCADA[SystemRealizationSCADA.iloc[:, 1].str.contains("ΣΥΝΟΛΙΚΟ ΦΟΡΤΙΟ") == True].index[0]
        Net_index = SystemRealizationSCADA[(SystemRealizationSCADA.iloc[:, 1].str.contains("NET LOAD") == True) | (SystemRealizationSCADA.iloc[:, 1].str.contains("ΚΑΘΑΡΟ ΦΟΡΤΙΟ") == True)].index[0]
        GrossLoad = []
        getScadaDataFromRange(SystemRealizationSCADA, Gross_index, Gross_index, Date, GrossLoad)
        NetLoad = []
        getScadaDataFromRange(SystemRealizationSCADA, Net_index, Net_index, Date, NetLoad)

        #handleVirtualUnits(results, file, Date, HVLoad)
        for i in range(len(GrossLoad)):
            for period in range(1, 5): 
                results.append({
                        'DispatchDay': GrossLoad[i][0].strftime('%Y-%m-%d'),
                        'DispatchPeriod': (int(GrossLoad[i][1]) - 1) * 4 + period,
                        'GrossValue': float(GrossLoad[i][3]) if not math.isnan(GrossLoad[i][3]) else 0,
                        'NetValue': float(NetLoad[i][3]) if not math.isnan(NetLoad[i][3]) else 0,
                        'ZoneId': 1,
                        'FileId': file.Id,
                        'Version': file.PublicationDate
                    })
        return envelope.getSuccessEnvelope(getParsingMetadataPayload(results, getMetadata(True, file.Id)))
    except Exception as e:
        message = f'Failed to parse {path}: {e}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

def ReadSCADAFile(path):
    SystemRealizationSCADA = pd.read_excel(path, sheet_name=1, skiprows=2, verbose = True)
    if len(SystemRealizationSCADA.columns) == 3:
        SystemRealizationSCADA = pd.read_excel(path, sheet_name=0, skiprows=2, verbose = True)

    return SystemRealizationSCADA

def getRealTimeScadaRes(dateFrom: str, dateTo : str):
    files = getFileUrlsFromApi("RealTimeSCADARES", dateFrom, dateTo) 

    for file in files:
        yield getScadaSystemLoadRealizationFromFile(file)

def getRealTimeScadaResFromFile(file: CustomFile):
    results = []

    path = getFileNameFromInfo(file)
    if not os.path.exists(path):
        message = f'{path} does not exist'
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        RealTimeSCADARES = pd.read_excel(path, skiprows = 1)
    except Exception as e:
        message = f'Failed to read {path}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        Date = RealTimeSCADARES['Date'][0]
        Version = file.PublicationDate
        date_to_datetime = datetime.strptime(Date, '%d-%m-%Y')

        date_to_datetime_pd = pd.Timestamp(date_to_datetime + timedelta(days=1) - timedelta(hours=1))
        hours = getTimeDifferncePandas(date_to_datetime_pd)
        dispatchDate = EETtoCET(date_to_datetime)

        for i in range(1, hours + 1):
            data = RealTimeSCADARES.values[0]
            if i >= len(data):
                break

            for period in range(1, 5): 
                dispatchPeriod = getQuarterlyDispatchPeriodFromHourString(dispatchDate.strftime('%H:%M'))
                print(dispatchPeriod)
                print(dispatchDate)
                results.append({
                        'DispatchDay': dispatchDate.date(),
                        'DispatchPeriod': dispatchPeriod,
                        'ZoneId': 1,
                        'FileId': file.Id,
                        'Value': float(data[i]),
                        'Version': Version
                    })
                dispatchDate += timedelta(minutes=15)
    
        return envelope.getSuccessEnvelope(getParsingMetadataPayload(results, getMetadata(True, file.Id)))
    except Exception as e:
        message = f'Failed to parse {path}: {e}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

def getRESActualMVInjections(dateFrom: datetime, dateTo: datetime):
    files = getFileUrlsFromApi("RESMV", dateFrom, dateTo) 

    for file in files:
        yield getRESActualMVInjectionsFromFile(file)

def getRESActualMVInjectionsFromFile(file: CustomFile):
    results = []

    path = getFileNameFromInfo(file)
    if not os.path.exists(path):
        message = f'{path} does not exist'
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        df = pd.read_excel(path)
    except Exception as e:
        message = f'Failed to read {path}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:    
        date = file.TargetDateFrom

        BaEnergy = df['B/A ΕΝΕΡΓΕΙΑ KWh'].values
        BaPower = df['B/A ΕΓΚ ΙΣΧΥΣ MW'].values
        SmallHydroEnergy = df['ΜΥΗΣ ΕΝΕΡΓΕΙΑ KWh'].values
        SmallHydroPower = df['ΜΥΗΣ ΕΓΚ ΙΣΧΥΣ MW'].values
        CoprodEnergy = df['ΣΗΘΥΑ ΕΝΕΡΓΕΙΑ KWh'].values
        CoprodPower = df['ΣΗΘΥΑ ΕΓΚ ΙΣΧΥΣ MW'].values
        PVEnergy = df['Φ/Β ΕΝΕΡΓΕΙΑ KWh'].values
        PVPower = df['Φ/Β ΕΓΚ ΙΣΧΥΣ MW'].values
        time_column = df['ΩΡΑ'].dropna()
        dispatchDate = EETtoCET(datetime.strptime(df['ΗΜΕΡΑ'][0],'%Y%m%d'))
        
        for t in range(1, len(time_column)+ 1 ):
             # Data contains info for yesterday

            for period in range(1, 5): 
                dispatchPeriod = getQuarterlyDispatchPeriodFromHourString(dispatchDate.strftime('%H:%M'))
                results.append({ 'DispatchDay': dispatchDate.date(), 
                                    'DispatchPeriod': dispatchPeriod ,
                                    'Version': file.PublicationDate,
                                    'FileId': file.Id,
                                    'ZoneId': 1,
                                    'BaEnergy': np.round(float(BaEnergy[t-1] / 1000), 3),
                                    'BaPower': np.round(float(BaPower[t-1]), 3),
                                    'SmallHydroEnergy': np.round(float(SmallHydroEnergy[t-1]) / 1000, 3),
                                    'SmallHydroPower': np.round(float(SmallHydroPower[t-1]), 3),
                                    'CoprodEnergy': np.round(float(CoprodEnergy[t-1]) / 1000, 3),
                                    'CoprodPower': np.round(float(CoprodPower[t-1]), 3),
                                    'PvEnergy': np.round(float(PVEnergy[t-1]) / 1000, 3),
                                    'PvPower': np.round(float(PVPower[t-1]), 3)
                                })
                dispatchDate += timedelta(minutes=15)

        return envelope.getSuccessEnvelope(getParsingMetadataPayload(results, getMetadata(True, file.Id)))
    except Exception as e:
        message = f'Failed to parse {path}: {e}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

def getLtNominations(dateFrom: datetime, dateTo: datetime):
    files = getFileUrlsFromApi("LTPTRsNominationsSummary", dateFrom, dateTo) 

    for file in files:
        yield getLtNominationsFromFile(file)

def getLtNominationsFromFile(file: CustomFile):
    results = []

    path = getFileNameFromInfo(file)
    if not os.path.exists(path):
        message = f'{path} does not exist'
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        df = pd.read_excel(path, skiprows = 1)
        if 'CET Hour' not in df.columns:
            df = pd.read_excel(path, skiprows = 2)
    except Exception as e:
        message = f'Failed to read {path}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        date = file.TargetDateFrom
        
        country_column = df.columns.get_loc("CET Hour")
        sum_indexes = df[df.iloc[:, country_column].str.contains("Sum") == True].index

        time_column = df.iloc[0,3:]
        index_import = sum_indexes[0]
        index_export = sum_indexes[1]

        countries_number = index_export - index_import

        for country_index in range(0, countries_number - 1):
            country = df.values[country_index, country_column]

            for t in range(1, len(time_column) + 1):
                if country_column + t >= len(df.values[country_index]):
                    break
                imports = df.values[country_index, country_column + t]
                exports = df.values[country_index + countries_number, country_column + t]

                for period in range(1, 5): 
                    results.append({ 
                        'DispatchDay': date, 
                        'DispatchPeriod': (t - 1) * 4 + period,
                        'Version': file.PublicationDate,
                        'CountryName': country,
                        'FileId': file.Id,
                        'Import': float(imports),
                        'Export': float(exports)
                    })

        return envelope.getSuccessEnvelope(getParsingMetadataPayload(results, getMetadata(True, file.Id)))
    except Exception as e:
        message = f'Failed to parse {path}: {e}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

def getActivatedEnergyAndSettlementPrice(dateFrom: str, dateTo : str):
    results = []

    files = getFileUrlsFromApi("IMBABE", dateFrom, dateTo) 

    for file in files:
        df = pd.read_excel(file["Url"])
        for index, row in df.iterrows():
            StartDate = row["STARTDATE"] if type(row["STARTDATE"]) == pd.Timestamp else datetime.strptime(row["STARTDATE"], '%m/%d/%Y %H:%M')
            
            TotalActivatedBalancingEnergyUP = float(row['Total Activated Balancing Energy UP (MWh)']) if not math.isnan(row['Total Activated Balancing Energy UP (MWh)']) else 0
            TotalActivatedBalancingEnergyDown = float(row['Total Activated Balancing Energy UP (MWh)']) if not math.isnan(row['Total Activated Balancing Energy UP (MWh)']) else 0
            TotalActivatedBalancingEnergy = float(row['Total Activated Balancing Energy (MWh)']) if not math.isnan(row['Total Activated Balancing Energy (MWh)']) else 0
            ImbalancePrice = float(row['Imbalance Price  (€/MWh)']) if not math.isnan(row['Imbalance Price  (€/MWh)']) else 0
            mFRRPriceUP = float(row['mFRR Price UP (€/MWh)']) if not math.isnan(row['mFRR Price UP (€/MWh)']) else 0
            mFRRPriceDown = float(row['mFRR Price Down (€/MWh)']) if not math.isnan(row['mFRR Price Down (€/MWh)']) else 0
            UpliftAccount1 = float(row['Uplift Account 1 (€/MWh)']) if not math.isnan(row['Uplift Account 1 (€/MWh)']) else 0
            UpliftAccount2 = float(row['Uplift Account 2 (€/MWh)']) if not math.isnan(row['Uplift Account 2 (€/MWh)']) else 0
            UpliftAccount3 = float(row['Uplift Account 3 (€/MWh)']) if not math.isnan(row['Uplift Account 3 (€/MWh)']) else 0
            BalancingEnergyIDEV = float(row['Balancing Energy IDEV (MWh)']) if 'Balancing Energy IDEV (MWh)' in row and not math.isnan(row['Balancing Energy IDEV (MWh)']) else 0
            BalancingEnergyUDEV = float(row['Balancing Energy UDEV (MWh)']) if 'Balancing Energy UDEV (MWh)' in row and not math.isnan(row['Balancing Energy UDEV (MWh)']) else 0

            for period in range(1, 5): 
                results.append({
                        'DispatchDay': StartDate.strftime("%Y-%m-%d"),
                        'DispatchPeriod': (int(getQuarterlyDispatchPeriodFromHourString(StartDate.strftime("%H:%M"))) - 1) * 4 + period,
                        'TotalActivatedBalancingEnergyUP': float(TotalActivatedBalancingEnergyUP),
                        'TotalActivatedBalancingEnergyDown': float(TotalActivatedBalancingEnergyDown),
                        'TotalActivatedBalancingEnergy': float(TotalActivatedBalancingEnergy),
                        'ImbalancePrice': float(ImbalancePrice),
                        'mFRRPriceUP': float(mFRRPriceUP),
                        'mFRRPriceDown': float(mFRRPriceDown),
                        'UpliftAccount1': float(UpliftAccount1),
                        'UpliftAccount2': float(UpliftAccount2),
                        'UpliftAccount3': float(UpliftAccount3),
                        'BalancingEnergyIDEV': float(BalancingEnergyIDEV),
                        'BalancingEnergyUDEV': float(BalancingEnergyUDEV),
                        'Version': datetime.strptime(file['PublicationDate'], '%d.%m.%Y %H:%M')
                    })

    return results

def getReservoirFillingRate(dateFrom: datetime, dateTo: datetime):
    files = getFileUrlsFromApi("ReservoirFillingRate", dateFrom, dateTo) 

    for file in files:
        yield getReservoirFillingRateFromFile(file)

def getReservoirFillingRateFromFile(file: CustomFile):
    path = getFileNameFromInfo(file)
    if not os.path.exists(path):
        message = f'{path} does not exist'
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        df = pd.read_excel(path, skiprows = 2)
    except Exception as e:
        message = f'Failed to read {path}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    results = []

    try:
        date = file.TargetDateFrom
        version = file.PublicationDate

        for i in range(len(df.index)):
            results.append({
                'DispatchDay': date, 
                'FileId': file.Id,
                'Version': version,
                'EntityName': df.iloc[i,1], 
                'FillingRate': float(df.iloc[i,2]) if not math.isnan(df.iloc[i,2]) else None,
                'MaxDailyEnergy': float(df.iloc[i,3]) if not math.isnan(df.iloc[i,3]) else None 
            })

        return envelope.getSuccessEnvelope(getParsingMetadataPayload(results, getMetadata(True, file.Id)))
    except Exception as e:
        message = f'Failed to parse {path}: {e}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)


def getBalancingEnergyProductsFromFile(file: CustomFile):
    path = getFileNameFromInfo(file)
    if not os.path.exists(path):
        message = f'{path} does not exist'
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)
    



    try:
        file_type = file.Url.split('.')[-1].lower()
        wb, sheet, interface = readWorkbook(path)
    except Exception as e:
        message = f'Failed to read {path}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    results = []

    try:
        expected_columns = [
            'STARTDATE',
            'ENDDATE',
            'Total mFRR UP (MWh)',
            'Total DA mFRR UP (MWh)',
            'Total mFRR DN (MWh)',
            'Total DA mFRR DN (MWh)',
            'Total aFRR UP (MWh)',
            'Total aFRR DN (MWh)'
        ]
        header_row = sheet[1] if file_type == 'xlsx' else sheet[0]
        header_values = [cell.value for cell in header_row]

        if header_values != expected_columns:
            message = f'Invalid column names in the sheet'
            raise Exception(message)

        version = file.PublicationDate

        timestamps = pd.read_excel(path)['STARTDATE']
        time_to_pd = pd.to_datetime(timestamps)
        time_infered = time_to_pd.dt.tz_localize('Europe/Berlin',ambiguous="infer")

        results = getAdmieEnergyProductDataFromFileType(sheet, time_infered,file_type, file, version, results, interface=interface)


        return envelope.getSuccessEnvelope(getParsingMetadataPayload(results, getMetadata(True, file.Id)))
    except Exception as e:
        message = f'Failed to parse {path}: {e}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

def getBalancingCapacityProductsFromFile(file: CustomFile):
    path = getFileNameFromInfo(file)
    if not os.path.exists(path):
        message = f'{path} does not exist'
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        file_type = file.Url.split('.')[-1].lower()
        wb, sheet, interface = readWorkbook(path)
    except Exception as e:
        message = f'Failed to read {path}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    results = []

    try:
        expected_columns = [
            'STARTDATE',
            'ENDDATE',
            'FCRQ_UP_ALL (MW)',
            'FCRQ_DN_ALL (MW)',
            'aFRRQ_UP_ALL (MW)',
            'aFRRQ_DN_ALL (MW)',
            'mFRRQ_UP_ALL (MW)',
            'mFRRQ_DN_ALL (MW)',
        ]

        header_row = sheet[1] if file_type == 'xlsx' else sheet[0]
        header_values = [cell.value for cell in header_row]

        if header_values != expected_columns:
            message = f'Invalid column names in the sheet'
            raise Exception(message)
        
        version = file.PublicationDate

        timestamps = pd.read_excel(path)['STARTDATE']
        time_to_pd = pd.to_datetime(timestamps)
        time_infered = time_to_pd.dt.tz_localize('Europe/Berlin',ambiguous="infer")

        results = getAdmieCapacityProductDataFromFileType(sheet, time_infered,file_type, file, version, results, interface=interface)

        return envelope.getSuccessEnvelope(getParsingMetadataPayload(results, getMetadata(True, file.Id)))
    except Exception as e:
        message = f'Failed to parse {path}: {e}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

def getAvailableTransferCapacity(dateFrom: datetime, dateTo: datetime):
    files = getFileUrlsFromApi("DailyAuctionsSpecificationsATC", dateFrom, dateTo) 

    for file in files:
        yield getAvailableTransferCapacityFromFile(file)

def getAvailableTransferCapacityFromFile(file: CustomFile):
    path = getFileNameFromInfo(file)
    if not os.path.exists(path):
        message = f'{path} does not exist'
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        xls = pd.ExcelFile(path)

        df = pd.read_excel(xls)
    except Exception as e:
        message = f'Failed to read {path}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    results = []
    try:
        df.dropna(how='all', axis=1, inplace=True)
        export = 'EXP' in df.columns[0]

        df = pd.read_excel(xls, skiprows=[0,1])
        df.dropna(how='all', axis=1, inplace=True)

        date = file.TargetDateFrom
        dispatch_hours = df.iloc[:, 0].values
        for country_index in range(1, len(df.columns)):
            country = df.columns[country_index]

            for t in range(1, len(dispatch_hours) + 1):

                if not checkStringToInt(dispatch_hours[t - 1]):
                    continue

                for j in range(1,5):
                    results.append({ 
                        'DispatchDay': date, 
                        'DispatchPeriod': 4*(int(dispatch_hours[t-1])-1) + j,
                        'Version': file.PublicationDate,
                        'FileId': file.Id,
                        'CountryNameFrom': 'GR' if export else country,
                        'CountryNameTo': 'GR' if not export else country,
                        'Value': float(df.values[t - 1, country_index])
                    })

        return envelope.getSuccessEnvelope(getParsingMetadataPayload(results, getMetadata(True, file.Id)))
    except Exception as e:
        message = f'Failed to parse {path}: {e}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

def getIspEnergySurplusFromFile(file: CustomFile) -> Envelope[List[VirtualFileMetadataPayload[List[EnergySurplusData]]]]:
    results = []
    version = file.PublicationDate

    path = getFileNameFromInfo(file)
    if not os.path.exists(path):
        message = f'{path} does not exist'
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)

    try:
        ISP_DataFrame = pd.read_excel(path, skiprows = 1, sheet_name=0)
    except Exception as e:
        message = f'Failed to read {path}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)
    
    try:
        first_column = ISP_DataFrame.columns[0]

        index_energy_surplus = ISP_DataFrame.index[ISP_DataFrame[first_column] == 'Energy Surplus'].values[-1]

        ISPEnergySurplus = []
        
        getISPResultsDataFromRange(ISP_DataFrame, index_energy_surplus, index_energy_surplus + 1, None,ISPEnergySurplus )

        dispatch_day = pd.Series([row[0] for row in ISPEnergySurplus]).unique()[0]

        # handle market granularity change on 1st Oct 2025
        market_granularity = 30 if pd.Timestamp(dispatch_day) < pd.Timestamp('2025-10-01') else 15

        raw_timestamps  = [datetime.strptime(row[0], '%Y-%m-%d') + timedelta(minutes= (row[1] - 1) * market_granularity)for row in ISPEnergySurplus]

        #get unique date from IspEnergySurplus
        timestamps = pd.Series(pd.date_range(
            start=pd.Timestamp(dispatch_day, tz='Europe/Berlin'),
            periods=len(ISPEnergySurplus),
            freq=f'{market_granularity}min'
        )).dt.tz_convert('UTC')

        for number, row in enumerate(ISPEnergySurplus):
            results.append(EnergySurplusData(**{
                'Timestamp': timestamps[number],
                'ZoneId': 1,
                'Value': row[3],
                'Version': version,
                "DispatchDay":raw_timestamps[number].strftime('%Y-%m-%d'),
            }))

        return envelope.getSuccessEnvelope(getParsingMetadataPayload(results, getMetadata(True, file.Id)))
    except Exception as e:
        message = f'Failed to parse {path}: {e}'
        logException(e)
        return envelope.getFailedEnvelope(getParsingMetadataPayload([], getMetadata(False, file.Id, message)), None)